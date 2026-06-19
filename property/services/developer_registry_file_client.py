"""Local file client for one-time developer registry imports."""

import csv
import json
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .developer_registry_client import (
    DeveloperRegistryClientError,
    DeveloperRegistrySourceItem,
    extract_rows_from_payload,
)


DEFAULT_FILE_SOURCE_REGION_CODE = 'file'
SUPPORTED_SOURCE_FILE_EXTENSIONS = ('.csv', '.json', '.xlsx')
CSV_DELIMITERS = ',;\t'
HEADER_ALIAS_MAP = {
    'actual address': 'actualAddress',
    'actualaddress': 'actualAddress',
    'address actual': 'actualAddress',
    'addressactual': 'actualAddress',
    'address legal': 'legalAddress',
    'addresslegal': 'legalAddress',
    'company group': 'companyGroupName',
    'company group name': 'companyGroupName',
    'companygroupname': 'companyGroupName',
    'developer': 'name',
    'developer name': 'name',
    'developername': 'name',
    'fact address': 'actualAddress',
    'factaddress': 'actualAddress',
    'factual address': 'actualAddress',
    'factualaddress': 'actualAddress',
    'group': 'companyGroupName',
    'group name': 'companyGroupName',
    'groupname': 'companyGroupName',
    'holding': 'companyGroupName',
    'holding name': 'companyGroupName',
    'holdingname': 'companyGroupName',
    'inn': 'inn',
    'jur address': 'legalAddress',
    'juraddress': 'legalAddress',
    'juridical address': 'legalAddress',
    'juridicaladdress': 'legalAddress',
    'kpp': 'kpp',
    'legal address': 'legalAddress',
    'legaladdress': 'legalAddress',
    'name': 'name',
    'ogrn': 'ogrn',
    'org name': 'name',
    'organization name': 'name',
    'organizationname': 'name',
    'short name': 'name',
    'shortname': 'name',
    'адрес регистрации': 'legalAddress',
    'адрес фактический': 'actualAddress',
    'адрес юридический': 'legalAddress',
    'гк': 'companyGroupName',
    'группа': 'companyGroupName',
    'группа компаний': 'companyGroupName',
    'застройщик': 'name',
    'инн': 'inn',
    'кпп': 'kpp',
    'место нахождения': 'legalAddress',
    'наименование': 'name',
    'наименование гк': 'companyGroupName',
    'наименование застройщика': 'name',
    'название': 'name',
    'название гк': 'companyGroupName',
    'название застройщика': 'name',
    'огрн': 'ogrn',
    'организация': 'name',
    'почтовый адрес': 'actualAddress',
    'факт адрес': 'actualAddress',
    'факт. адрес': 'actualAddress',
    'фактический адрес': 'actualAddress',
    'холдинг': 'companyGroupName',
    'юр адрес': 'legalAddress',
    'юр. адрес': 'legalAddress',
    'юридический адрес': 'legalAddress',
    'юридическое лицо': 'name',
}


class DeveloperRegistryFileError(DeveloperRegistryClientError):
    """Raised when a local developer registry file cannot be read."""


class FileDeveloperRegistryClient:
    """Read developer registry rows from a local CSV, XLSX, or JSON file."""

    def __init__(
        self,
        source_file_path,
        source_region_code=DEFAULT_FILE_SOURCE_REGION_CODE,
    ):
        """Store file import settings."""
        self.source_file_path = Path(source_file_path)
        self.source_region_code = str(source_region_code)

    def fetch_developers(self, region_codes=None, limit=None):
        """Yield developer rows from the configured local file."""
        yielded_count = 0

        for payload in load_developer_registry_file(self.source_file_path):
            yield DeveloperRegistrySourceItem(
                region_code=self.source_region_code,
                payload=payload,
            )
            yielded_count += 1
            if limit and yielded_count >= limit:
                return


def load_developer_registry_file(source_file_path):
    """Load registry rows from a supported local source file."""
    path = Path(source_file_path)
    if not path.exists():
        raise DeveloperRegistryFileError(
            f'Developer registry source file does not exist: {path}'
        )
    if not path.is_file():
        raise DeveloperRegistryFileError(
            f'Developer registry source path is not a file: {path}'
        )

    extension = path.suffix.casefold()
    if extension == '.csv':
        return load_csv_developer_registry_file(path)
    if extension == '.json':
        return load_json_developer_registry_file(path)
    if extension == '.xlsx':
        return load_xlsx_developer_registry_file(path)

    supported_extensions = ', '.join(SUPPORTED_SOURCE_FILE_EXTENSIONS)
    raise DeveloperRegistryFileError(
        (
            f'Unsupported developer registry source file extension '
            f'{extension}. Supported extensions: {supported_extensions}.'
        )
    )


def load_csv_developer_registry_file(path):
    """Load registry rows from a UTF-8 CSV file."""
    try:
        with path.open('r', encoding='utf-8-sig', newline='') as source_file:
            sample = source_file.read(4096)
            source_file.seek(0)
            reader = csv.DictReader(
                source_file,
                dialect=detect_csv_dialect(sample),
            )
            if not reader.fieldnames:
                return []
            return [
                normalize_source_row(row, row_number=row_number)
                for row_number, row in enumerate(reader, start=2)
                if row_has_values(row.values())
            ]
    except (OSError, UnicodeDecodeError, csv.Error) as exception:
        raise DeveloperRegistryFileError(
            f'Cannot read developer registry CSV file {path}: {exception}'
        ) from exception


def load_json_developer_registry_file(path):
    """Load registry rows from a UTF-8 JSON file."""
    try:
        payload = json.loads(path.read_text(encoding='utf-8-sig'))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exception:
        raise DeveloperRegistryFileError(
            f'Cannot read developer registry JSON file {path}: {exception}'
        ) from exception

    rows = extract_rows_from_payload(payload)
    if not rows and isinstance(payload, dict):
        rows = [payload]

    return [
        normalize_source_row(row, row_number=row_number)
        for row_number, row in enumerate(rows, start=1)
        if isinstance(row, dict) and row_has_values(row.values())
    ]


def load_xlsx_developer_registry_file(path):
    """Load registry rows from the first worksheet of an XLSX file."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (
        OSError,
        InvalidFileException,
        BadZipFile,
        ValueError,
    ) as exception:
        raise DeveloperRegistryFileError(
            f'Cannot read developer registry XLSX file {path}: {exception}'
        ) from exception

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []

        normalized_headers = [
            normalize_source_header(header) for header in headers
        ]
        source_rows = []
        for row_number, row in enumerate(rows, start=2):
            if not row_has_values(row):
                continue

            source_rows.append(
                normalize_source_row(
                    {
                        header: value
                        for header, value in zip(normalized_headers, row)
                        if header
                    },
                    row_number=row_number,
                    headers_already_normalized=True,
                )
            )
        return source_rows
    finally:
        workbook.close()


def detect_csv_dialect(sample):
    """Detect a CSV dialect, defaulting to comma-separated files."""
    try:
        return csv.Sniffer().sniff(sample, delimiters=CSV_DELIMITERS)
    except csv.Error:
        return csv.excel


def normalize_source_row(
    row,
    row_number=None,
    headers_already_normalized=False,
):
    """Normalize source row headers while preserving source values."""
    normalized_row = {}
    for key, value in row.items():
        normalized_key = (
            key if headers_already_normalized else normalize_source_header(key)
        )
        if not normalized_key:
            continue
        normalized_row[normalized_key] = normalize_source_value(value)

    if row_number is not None:
        normalized_row['_sourceRowNumber'] = row_number
    return normalized_row


def normalize_source_header(header):
    """Map common Russian and English file headers to importer aliases."""
    if header in (None, ''):
        return ''

    header_text = str(header).replace('\ufeff', '').replace('\xa0', ' ')
    normalized_header = ' '.join(
        header_text.strip().casefold().replace('ё', 'е').split()
    )
    return HEADER_ALIAS_MAP.get(normalized_header, header_text.strip())


def normalize_source_value(value):
    """Normalize a source cell value without changing its domain meaning."""
    if isinstance(value, dict):
        return normalize_source_row(value)
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def row_has_values(values):
    """Return whether a source row has at least one non-empty value."""
    return any(value not in (None, '') for value in values)
