# mortgage/views.py
import decimal

import openpyxl
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from location.models import City, District
from property.models import (
    ApartmentDecoration,
    ApartmentLayout,
    Developer,
    Property,
    RealEstateComplex,
    RealEstateComplexBuilding,
)
from trench_mortgage.views import (
    _build_trench_input_rows,
    _calculate_trench_mortgage,
    _export_trench_excel,
    _format_payment_schedule as _format_trench_payment_schedule,
    _format_result as _format_trench_result,
    _parse_trench_inputs,
    _prepare_mortgage_data as _prepare_trench_mortgage_data,
    _resolve_default_rate as _resolve_trench_default_rate,
    _resolve_trench_count,
    _save_trench_calculation,
)

from .forms import MortgageForm
from .models import MortgageCalculation
from .mortgage_calculator import MortgageCalculator
from .utils import (
    apply_calculation_filters,
    apply_calculation_sort,
    annotate_calculation_table_values,
    build_calculation_table_headers,
    format_currency,
    get_calculation_city_choices,
    get_calculation_filters,
    get_calculation_sort,
)


def _get_target_customer(request):
    customer_id = request.POST.get('customer') or request.GET.get('customer')
    if not customer_id or not request.user.is_authenticated:
        return None

    from customer.models import Customer

    return get_object_or_404(Customer, pk=customer_id, user=request.user)


def _attach_calculation_to_customer(customer, calculation):
    if customer is None:
        return

    from customer.models import CustomerCalculation

    CustomerCalculation.objects.get_or_create(
        customer=customer,
        calculation=calculation,
    )


def _normalize_discount_markup_values(cleaned_data):
    """Р’РѕР·РІСЂР°С‰Р°РµС‚ РїСЂРѕС†РµРЅС‚, СЂСѓР±Р»Рё Рё РёС‚РѕРіРѕРІСѓСЋ СЃС‚РѕРёРјРѕСЃС‚СЊ РїРѕСЃР»Рµ РєРѕСЂСЂРµРєС‚РёСЂРѕРІРєРё."""
    property_cost = float(cleaned_data['PROPERTY_COST'])
    discount_markup_percent = float(
        cleaned_data.get('DISCOUNT_MARKUP_VALUE', 0) or 0
    )
    discount_markup_rubles = float(
        cleaned_data.get('DISCOUNT_MARKUP_RUBLES', 0) or 0
    )
    discount_markup_source = cleaned_data.get('DISCOUNT_MARKUP_SOURCE')

    if discount_markup_source == 'rubles':
        if property_cost > 0:
            discount_markup_percent = (
                discount_markup_rubles / property_cost
            ) * 100
        else:
            discount_markup_percent = 0
    else:
        discount_markup_rubles = (
            property_cost * discount_markup_percent / 100
        )

    if cleaned_data['DISCOUNT_MARKUP_TYPE'] == 'discount':
        final_property_cost = property_cost - discount_markup_rubles
    else:
        final_property_cost = property_cost + discount_markup_rubles

    return (
        discount_markup_percent,
        discount_markup_rubles,
        final_property_cost,
    )


def _normalize_initial_payment_values(cleaned_data, final_property_cost):
    """Р’РѕР·РІСЂР°С‰Р°РµС‚ СЃРѕРіР»Р°СЃРѕРІР°РЅРЅС‹Рµ Р·РЅР°С‡РµРЅРёСЏ РїРµСЂРІРѕРЅР°С‡Р°Р»СЊРЅРѕРіРѕ РІР·РЅРѕСЃР°."""
    initial_payment_percent = float(
        cleaned_data.get('INITIAL_PAYMENT_PERCENT', 0) or 0
    )
    initial_payment_rubles = float(
        cleaned_data.get('INITIAL_PAYMENT_RUBLES', 0) or 0
    )
    initial_payment_source = cleaned_data.get('INITIAL_PAYMENT_SOURCE')

    if initial_payment_source == 'rubles':
        initial_payment_percent = (
            initial_payment_rubles / final_property_cost * 100
            if final_property_cost > 0
            else 0
        )
    else:
        initial_payment_rubles = (
            final_property_cost * initial_payment_percent / 100
        )

    return initial_payment_percent, initial_payment_rubles


def _populate_market_report_context(
    context,
    data,
    final_property_cost,
    initial_payment_percent,
):
    """Add a formatted market mortgage report to the template context."""
    if (
        data.get('HAS_GRACE_PERIOD') == 'yes'
        and (
            data.get('GRACE_PERIOD_TERM') in (None, '')
            or data.get('GRACE_PERIOD_RATE') in (None, '')
        )
    ):
        return False

    calculator = MortgageCalculator(
        property_cost=float(final_property_cost),
        initial_payment_percent=float(initial_payment_percent),
        initial_payment_date=data['INITIAL_PAYMENT_DATE'],
        mortgage_term=int(data['MORTGAGE_TERM']),
        annual_rate=float(data['ANNUAL_RATE']),
        has_grace_period=data['HAS_GRACE_PERIOD'] == 'yes',
        grace_period_term=int(data['GRACE_PERIOD_TERM'] or 0),
        grace_period_rate=float(data['GRACE_PERIOD_RATE'] or 0),
    )
    result = calculator.calculate()

    formatted_result = {}
    for key, value in result.items():
        if key in ['grace_payments_count', 'main_payments_count']:
            formatted_result[key] = int(value) if value else 0
        elif isinstance(value, (int, float, decimal.Decimal)):
            formatted_result[key] = format_currency(value)
        else:
            formatted_result[key] = value

    payment_schedule = calculator.get_payment_schedule()
    for payment in payment_schedule:
        for key in [
            'payment_amount',
            'interest_amount',
            'principal_amount',
            'remaining_debt',
        ]:
            if key in payment:
                payment[key] = format_currency(payment[key])

    context['result'] = formatted_result
    context['market_result'] = formatted_result
    context['has_grace_period'] = data['HAS_GRACE_PERIOD'] == 'yes'
    context['payment_schedule'] = payment_schedule
    context['market_payment_schedule'] = payment_schedule
    return result


def _populate_trench_report_context(
    context,
    request,
    data,
    property_obj,
    report_errors=False,
):
    """Add a formatted trench mortgage report to the template context."""
    trench_mortgage_data, prep_errors = (
        _prepare_trench_mortgage_data(data)
    )
    trench_entries, input_rows, trench_errors = _parse_trench_inputs(
        post_data=request.POST,
        trench_count=trench_mortgage_data['trench_count'],
        loan_amount=trench_mortgage_data['total_loan_amount'],
        default_annual_rate=trench_mortgage_data['annual_rate'],
    )
    context['trench_input_rows'] = input_rows
    all_errors = prep_errors + trench_errors
    if all_errors:
        if report_errors:
            context['error_message'] = ' '.join(all_errors)
            context['active_calculation_type'] = 'trench'
        return None

    trench_calculation, calc_errors = _calculate_trench_mortgage(
        trench_mortgage_data, trench_entries
    )
    if calc_errors:
        if report_errors:
            context['error_message'] = ' '.join(calc_errors)
            context['active_calculation_type'] = 'trench'
        return None

    context['trench_result'] = _format_trench_result(trench_calculation)
    context['trench_payment_schedule'] = _format_trench_payment_schedule(
        trench_calculation['payment_schedule']
    )
    context['can_export_trench_result'] = True
    return trench_calculation


def _get_discount_markup_labels(discount_markup_type):
    """Р’РѕР·РІСЂР°С‰Р°РµС‚ РїРѕРґРїРёСЃРё РґР»СЏ РєРѕСЂСЂРµРєС‚РёСЂРѕРІРєРё С†РµРЅС‹ РІ РїСЂРѕС†РµРЅС‚Р°С… Рё СЂСѓР±Р»СЏС…."""
    if discount_markup_type == 'discount':
        return 'РЎРєРёРґРєР°, %', 'РЎРєРёРґРєР°, СЂСѓР±.'

    return 'РЈРґРѕСЂРѕР¶Р°РЅРёРµ, %', 'РЈРґРѕСЂРѕР¶Р°РЅРёРµ, СЂСѓР±.'


def _get_property_initial(property_obj):
    """Return calculator form initial data from an existing property."""
    real_estate_complex = property_obj.building.real_estate_complex
    district = real_estate_complex.district

    return {
        'OBJECT_CITY': district.city_id,
        'OBJECT_DISTRICT': district.pk,
        'OBJECT_DEVELOPER': real_estate_complex.developer_id,
        'OBJECT_COMPLEX': real_estate_complex.pk,
        'OBJECT_BUILDING': property_obj.building_id,
        'OBJECT_APARTMENT_NUMBER': property_obj.apartment_number,
        'OBJECT_AREA': property_obj.area,
        'OBJECT_LAYOUT': property_obj.layout_id,
        'OBJECT_FLOOR': property_obj.floor,
        'OBJECT_DECORATION': property_obj.decoration_id,
    }


def _get_property_payload(property_obj):
    """Return property data used by the calculator UI."""
    real_estate_complex = property_obj.building.real_estate_complex
    district = real_estate_complex.district

    return {
        'id': property_obj.pk,
        'property_cost': str(property_obj.property_cost),
        'city_id': district.city_id,
        'district_id': district.pk,
        'developer_id': real_estate_complex.developer_id,
        'complex_id': real_estate_complex.pk,
        'building_id': property_obj.building_id,
        'apartment_number': property_obj.apartment_number,
        'area': str(property_obj.area),
        'layout_id': property_obj.layout_id,
        'floor': property_obj.floor,
        'decoration_id': property_obj.decoration_id,
    }


def _get_property_form_data():
    """Build reusable selector data for the mortgage object block."""
    districts = District.objects.select_related('city').order_by('name')
    complexes = RealEstateComplex.objects.select_related(
        'developer',
        'district__city',
    ).order_by('name')
    buildings = RealEstateComplexBuilding.objects.select_related(
        'real_estate_complex'
    ).order_by('real_estate_complex__name', 'number')
    properties = Property.objects.select_related(
        'building',
        'building__real_estate_complex__developer',
        'building__real_estate_complex__district__city',
        'layout',
        'decoration',
    ).order_by('building_id', 'apartment_number')

    return {
        'cities': list(City.objects.order_by('name').values('id', 'name')),
        'districts': list(districts.values('id', 'name', 'city_id')),
        'complexes': list(
            complexes.values(
                'id',
                'name',
                'developer_id',
                'district_id',
                'district__city_id',
            )
        ),
        'buildings': list(
            buildings.values('id', 'number', 'real_estate_complex_id')
        ),
        'properties': [
            _get_property_payload(property_obj)
            for property_obj in properties
        ],
    }


def _get_selected_property_from_form_data(form_data):
    """Return a selected property from hidden id or apartment number."""
    selected_id = form_data.get('PROPERTY')
    if selected_id:
        return (
            Property.objects.select_related(
                'building__real_estate_complex__developer',
                'building__real_estate_complex__district__city',
                'building',
                'layout',
                'decoration',
            )
            .filter(id=selected_id)
            .first()
        )

    building_id = form_data.get('OBJECT_BUILDING')
    apartment_number = (form_data.get('OBJECT_APARTMENT_NUMBER') or '').strip()
    if not building_id or not apartment_number:
        return None

    return (
        Property.objects.select_related(
            'building__real_estate_complex__developer',
            'building__real_estate_complex__district__city',
            'building',
            'layout',
            'decoration',
        )
        .filter(
            building_id=building_id,
            apartment_number=apartment_number,
        )
        .order_by('pk')
        .first()
    )


def _create_manual_property(cleaned_data, property_cost):
    """Create a property from manually filled calculator object data."""
    return Property.objects.create(
        apartment_number=cleaned_data['OBJECT_APARTMENT_NUMBER'],
        building=cleaned_data['OBJECT_BUILDING'],
        decoration=cleaned_data['OBJECT_DECORATION'],
        layout=cleaned_data['OBJECT_LAYOUT'],
        area=cleaned_data['OBJECT_AREA'],
        floor=cleaned_data['OBJECT_FLOOR'],
        property_cost=decimal.Decimal(str(property_cost)),
    )


def _build_calculation(property_obj, data, result, values):
    """Build a saved mortgage calculation for a property-backed scenario."""
    return MortgageCalculation(
        property=property_obj,
        base_property_cost=decimal.Decimal(str(values['base_property_cost'])),
        initial_payment_percent=decimal.Decimal(
            str(values['initial_payment_percent'])
        ),
        initial_payment_date=data['INITIAL_PAYMENT_DATE'],
        mortgage_term=data['MORTGAGE_TERM'],
        annual_rate=decimal.Decimal(str(data['ANNUAL_RATE'])),
        has_grace_period=data['HAS_GRACE_PERIOD'] == 'yes',
        grace_period_term=data['GRACE_PERIOD_TERM'],
        grace_period_rate=decimal.Decimal(
            str(data['GRACE_PERIOD_RATE'] or 0)
        ),
        discount_markup_type=data['DISCOUNT_MARKUP_TYPE'],
        discount_markup_value=decimal.Decimal(
            str(values['discount_markup_value'])
        ),
        final_property_cost=decimal.Decimal(
            str(values['final_property_cost'])
        ),
        grace_payments_count=result['grace_payments_count'],
        grace_period_end_date=result['grace_period_end_date'],
        grace_monthly_payment=decimal.Decimal(
            str(result['grace_monthly_payment'])
        ),
        loan_after_grace=decimal.Decimal(str(result['loan_after_grace'])),
        main_payments_count=result['main_payments_count'],
        mortgage_end_date=result['mortgage_end_date'],
        main_monthly_payment=decimal.Decimal(
            str(result['main_monthly_payment'])
        ),
        total_loan_amount=decimal.Decimal(str(result['total_loan_amount'])),
        total_overpayment=decimal.Decimal(str(result['total_overpayment'])),
    )


def _get_sample_calculation(request):
    """Р’РѕР·РІСЂР°С‰Р°РµС‚ СЂР°СЃС‡РµС‚-РѕР±СЂР°Р·РµС† РґР»СЏ РїСЂРµРґР·Р°РїРѕР»РЅРµРЅРёСЏ С„РѕСЂРјС‹ РєР°Р»СЊРєСѓР»СЏС‚РѕСЂР°."""
    sample_calculation_id = (request.GET.get('sample') or '').strip()
    if not sample_calculation_id or not sample_calculation_id.isdecimal():
        return None

    return get_object_or_404(
        MortgageCalculation.objects.select_related(
            'property',
            'property__building',
            'property__building__real_estate_complex',
            'property__building__real_estate_complex__district',
            'property__layout',
            'property__decoration',
        ),
        pk=sample_calculation_id,
    )


def _get_calculation_form_initial(calculation):
    """Р¤РѕСЂРјРёСЂСѓРµС‚ initial-РґР°РЅРЅС‹Рµ С„РѕСЂРјС‹ РёР· СЃРѕС…СЂР°РЅРµРЅРЅРѕРіРѕ СЂР°СЃС‡РµС‚Р°."""
    discount_markup_rubles = (
        calculation.base_property_cost
        * calculation.discount_markup_value
        / decimal.Decimal('100')
    )
    initial_payment_rubles = calculation.initial_payment_amount
    grace_period_term = calculation.grace_period_term or 0
    has_grace_period = 'yes' if calculation.has_grace_period else 'no'

    initial = {
        'PROPERTY': calculation.property_id,
        'PROPERTY_COST': calculation.base_property_cost,
        'DISCOUNT_MARKUP_TYPE': calculation.discount_markup_type,
        'DISCOUNT_MARKUP_VALUE': calculation.discount_markup_value,
        'DISCOUNT_MARKUP_RUBLES': discount_markup_rubles,
        'DISCOUNT_MARKUP_SOURCE': 'percent',
        'INITIAL_PAYMENT_PERCENT': calculation.initial_payment_percent,
        'INITIAL_PAYMENT_RUBLES': initial_payment_rubles,
        'INITIAL_PAYMENT_SOURCE': 'percent',
        'INITIAL_PAYMENT_DATE': calculation.initial_payment_date,
        'MORTGAGE_TERM_YEARS': calculation.mortgage_term // 12,
        'MORTGAGE_TERM': calculation.mortgage_term,
        'ANNUAL_RATE': calculation.annual_rate,
        'HAS_GRACE_PERIOD': has_grace_period,
        'GRACE_PERIOD_TERM_YEARS': grace_period_term // 12,
        'GRACE_PERIOD_TERM': grace_period_term,
        'GRACE_PERIOD_RATE': calculation.grace_period_rate,
    }
    initial.update(_get_property_initial(calculation.property))
    return initial


def mortgage_calculator(request):
    # РРЅРёС†РёР°Р»РёР·Р°С†РёСЏ С„РѕСЂРјС‹
    """РћРїРёСЃР°РЅРёРµ РјРµС‚РѕРґР° mortgage_calculator.

    Р’С‹РїРѕР»РЅСЏРµС‚ РїСЂРёРєР»Р°РґРЅСѓСЋ РѕРїРµСЂР°С†РёСЋ С‚РµРєСѓС‰РµРіРѕ РјРѕРґСѓР»СЏ.

    РђСЂРіСѓРјРµРЅС‚С‹:
        request: Р’С…РѕРґРЅРѕР№ РїР°СЂР°РјРµС‚СЂ, РІР»РёСЏСЋС‰РёР№ РЅР° СЂР°Р±РѕС‚Сѓ РјРµС‚РѕРґР°.

    Р’РѕР·РІСЂР°С‰Р°РµС‚:
        Any: РўРёРї СЂРµР·СѓР»СЊС‚Р°С‚Р° РѕРїСЂРµРґРµР»СЏРµС‚СЃСЏ РІС‹Р·С‹РІР°СЋС‰РёРј РєРѕРґРѕРј.
    """
    target_customer = _get_target_customer(request)
    sample_calculation = (
        _get_sample_calculation(request)
        if request.method == 'GET'
        else None
    )

    if request.method == 'POST':
        form_data = request.POST.copy()
        submitted_calculation_type = (
            form_data.get('calculate') or form_data.get('export')
        )
        if submitted_calculation_type in ('market', 'trench'):
            form_data['CALCULATION_TYPE'] = submitted_calculation_type
        selected_property = _get_selected_property_from_form_data(form_data)
        if selected_property:
            form_data['PROPERTY'] = str(selected_property.pk)
            if not form_data.get('PROPERTY_COST'):
                form_data['PROPERTY_COST'] = str(
                    selected_property.property_cost
                )
            for field_name, value in _get_property_initial(
                selected_property
            ).items():
                if not form_data.get(field_name):
                    form_data[field_name] = str(value)
        mortgage_form = MortgageForm(form_data)
    elif sample_calculation is not None:
        mortgage_form = MortgageForm(
            initial=_get_calculation_form_initial(sample_calculation)
        )
    else:
        mortgage_form = MortgageForm()

    posted_calculation_type = (
        request.POST.get('calculate') or request.POST.get('export')
    )
    active_calculation_type = (
        posted_calculation_type
        if posted_calculation_type in ('market', 'trench')
        else (
            mortgage_form.data.get('CALCULATION_TYPE')
            if mortgage_form.is_bound
            else mortgage_form.initial.get('CALCULATION_TYPE', 'market')
        )
    )
    if active_calculation_type not in ('market', 'trench'):
        active_calculation_type = 'market'
    trench_count = _resolve_trench_count(mortgage_form)
    trench_default_rate = _resolve_trench_default_rate(mortgage_form)

    context = {
        'mortgage_form': mortgage_form,
        'target_customer': target_customer,
        'sample_calculation': sample_calculation,
        'property_form_data': _get_property_form_data(),
        'active_calculation_type': active_calculation_type,
        'trench_count': trench_count,
        'trench_input_rows': _build_trench_input_rows(
            trench_count=trench_count,
            post_data=mortgage_form.data if mortgage_form.is_bound else None,
            default_annual_rate=trench_default_rate,
        ),
    }

    if request.method == 'POST':
        if 'calculate' in request.POST:
            if mortgage_form.is_valid():
                # РџРѕР»СѓС‡Р°РµРј РґР°РЅРЅС‹Рµ РёР· С„РѕСЂРјС‹
                data = mortgage_form.cleaned_data
                selected_calculation_type = request.POST.get('calculate')
                if selected_calculation_type not in ('market', 'trench'):
                    selected_calculation_type = (
                        data.get('CALCULATION_TYPE') or 'market'
                    )
                data['CALCULATION_TYPE'] = selected_calculation_type
                context['active_calculation_type'] = selected_calculation_type

                # РџРѕР»СѓС‡Р°РµРј РІС‹Р±СЂР°РЅРЅС‹Р№ РѕР±СЉРµРєС‚ РЅРµРґРІРёР¶РёРјРѕСЃС‚Рё
                property_obj = data['PROPERTY']

                # РџРѕР»СѓС‡Р°РµРј Р±Р°Р·РѕРІСѓСЋ СЃС‚РѕРёРјРѕСЃС‚СЊ РёР· СЃРєСЂС‹С‚РѕРіРѕ РїРѕР»СЏ
                base_property_cost = float(data['PROPERTY_COST'])

                (
                    discount_markup_value,
                    discount_markup_rubles,
                    final_property_cost,
                ) = _normalize_discount_markup_values(data)
                (
                    initial_payment_percent,
                    initial_payment_rubles,
                ) = _normalize_initial_payment_values(
                    data, final_property_cost
                )

                # РЎРѕР·РґР°РµРј СЌРєР·РµРјРїР»СЏСЂ РєР°Р»СЊРєСѓР»СЏС‚РѕСЂР°.
                # Р’СЃРµ Р·РЅР°С‡РµРЅРёСЏ РїСЂРµРѕР±СЂР°Р·СѓРµРј Рє float.
                if data.get('CALCULATION_TYPE') == 'trench':
                    trench_mortgage_data, prep_errors = (
                        _prepare_trench_mortgage_data(data)
                    )
                    trench_entries, input_rows, trench_errors = (
                        _parse_trench_inputs(
                            post_data=request.POST,
                            trench_count=trench_mortgage_data[
                                'trench_count'
                            ],
                            loan_amount=trench_mortgage_data[
                                'total_loan_amount'
                            ],
                            default_annual_rate=trench_mortgage_data[
                                'annual_rate'
                            ],
                        )
                    )
                    context['trench_input_rows'] = input_rows
                    all_errors = prep_errors + trench_errors
                    if all_errors:
                        context['error_message'] = ' '.join(all_errors)
                        context['active_calculation_type'] = 'trench'
                        return render(
                            request,
                            'mortgage/mortgage_form.html',
                            context,
                        )

                    trench_calculation, calc_errors = (
                        _calculate_trench_mortgage(
                            trench_mortgage_data,
                            trench_entries,
                        )
                    )
                    if calc_errors:
                        context['error_message'] = ' '.join(calc_errors)
                        context['active_calculation_type'] = 'trench'
                        return render(
                            request,
                            'mortgage/mortgage_form.html',
                            context,
                        )

                    should_save_calculation = (
                        property_obj is not None
                        or mortgage_form.has_manual_property_data()
                    )
                    if should_save_calculation:
                        with transaction.atomic():
                            if property_obj is None:
                                property_obj = _create_manual_property(
                                    data,
                                    base_property_cost,
                                )
                            trench_calculation['property_obj'] = property_obj
                            _save_trench_calculation(trench_calculation)

                    context['trench_result'] = _format_trench_result(
                        trench_calculation
                    )
                    context['trench_payment_schedule'] = (
                        _format_trench_payment_schedule(
                            trench_calculation['payment_schedule']
                        )
                    )
                    context['can_export_trench_result'] = True
                    context['active_calculation_type'] = 'trench'
                    context['final_property_cost'] = format_currency(
                        final_property_cost
                    )
                    context['discount_markup_type'] = data[
                        'DISCOUNT_MARKUP_TYPE'
                    ]
                    context['discount_markup_value'] = discount_markup_value
                    context['discount_markup_rubles'] = discount_markup_rubles
                    context['selected_property'] = property_obj
                    context['initial_payment_percent'] = (
                        initial_payment_percent
                    )
                    context['initial_payment_rubles'] = initial_payment_rubles
                    _populate_market_report_context(
                        context,
                        data,
                        final_property_cost,
                        initial_payment_percent,
                    )
                    context['active_calculation_type'] = 'trench'
                    context['mortgage_form'] = mortgage_form
                    return render(
                        request,
                        'mortgage/mortgage_form.html',
                        context,
                    )

                calculator = MortgageCalculator(
                    property_cost=float(final_property_cost),
                    initial_payment_percent=float(initial_payment_percent),
                    initial_payment_date=data['INITIAL_PAYMENT_DATE'],
                    mortgage_term=int(data['MORTGAGE_TERM']),
                    annual_rate=float(data['ANNUAL_RATE']),
                    has_grace_period=data['HAS_GRACE_PERIOD'] == 'yes',
                    grace_period_term=int(data['GRACE_PERIOD_TERM'] or 0),
                    grace_period_rate=float(data['GRACE_PERIOD_RATE'] or 0),
                )

                # Р’С‹РїРѕР»РЅСЏРµРј СЂР°СЃС‡РµС‚
                result = calculator.calculate()

                # Р¤РѕСЂРјР°С‚РёСЂСѓРµРј С‡РёСЃР»РѕРІС‹Рµ Р·РЅР°С‡РµРЅРёСЏ РґР»СЏ РѕС‚РѕР±СЂР°Р¶РµРЅРёСЏ
                formatted_result = {}
                for key, value in result.items():
                    if key in ['grace_payments_count', 'main_payments_count']:
                        # Р¦РµР»С‹Рµ С‡РёСЃР»Р°
                        formatted_result[key] = int(value) if value else 0
                    elif isinstance(value, (int, float, decimal.Decimal)):
                        # Р”РµРЅРµР¶РЅС‹Рµ Р·РЅР°С‡РµРЅРёСЏ
                        formatted_result[key] = format_currency(value)
                    else:
                        formatted_result[key] = value

                # РџРѕР»СѓС‡Р°РµРј РіСЂР°С„РёРє РїР»Р°С‚РµР¶РµР№
                payment_schedule = calculator.get_payment_schedule()

                # Р¤РѕСЂРјР°С‚РёСЂСѓРµРј С‡РёСЃР»РѕРІС‹Рµ Р·РЅР°С‡РµРЅРёСЏ РІ РіСЂР°С„РёРєРµ РїР»Р°С‚РµР¶РµР№
                for payment in payment_schedule:
                    for key in [
                        'payment_amount',
                        'interest_amount',
                        'principal_amount',
                        'remaining_debt',
                    ]:
                        if key in payment:
                            payment[key] = format_currency(payment[key])

                calculation = None
                should_save_calculation = (
                    property_obj is not None
                    or mortgage_form.has_manual_property_data()
                )
                if should_save_calculation:
                    calculation_values = {
                        'base_property_cost': base_property_cost,
                        'initial_payment_percent': initial_payment_percent,
                        'discount_markup_value': discount_markup_value,
                        'final_property_cost': final_property_cost,
                    }
                    with transaction.atomic():
                        if property_obj is None:
                            property_obj = _create_manual_property(
                                data,
                                base_property_cost,
                            )
                        calculation = _build_calculation(
                            property_obj,
                            data,
                            result,
                            calculation_values,
                        )
                        calculation.save()
                        _attach_calculation_to_customer(
                            target_customer,
                            calculation,
                        )

                    if target_customer is not None:
                        messages.success(
                            request,
                            'Р Р°СЃС‡РµС‚ СЃРѕС…СЂР°РЅРµРЅ Рё РїСЂРёРІСЏР·Р°РЅ Рє РєР»РёРµРЅС‚Сѓ.',
                        )

                # РЎРѕС…СЂР°РЅСЏРµРј СЂР°СЃС‡РµС‚ РІ РєРѕРЅС‚РµРєСЃС‚
                context['result'] = formatted_result
                context['market_result'] = formatted_result
                context['has_grace_period'] = data['HAS_GRACE_PERIOD'] == 'yes'
                context['payment_schedule'] = payment_schedule
                context['market_payment_schedule'] = payment_schedule
                context['active_calculation_type'] = 'market'
                context['final_property_cost'] = format_currency(
                    final_property_cost
                )
                context['discount_markup_type'] = data['DISCOUNT_MARKUP_TYPE']
                context['discount_markup_value'] = discount_markup_value
                context['discount_markup_rubles'] = discount_markup_rubles
                context['selected_property'] = property_obj
                context['initial_payment_percent'] = initial_payment_percent
                context['initial_payment_rubles'] = initial_payment_rubles

                # РџРµСЂРµРґР°РµРј Р·Р°РїРѕР»РЅРµРЅРЅСѓСЋ С„РѕСЂРјСѓ РІ РєРѕРЅС‚РµРєСЃС‚
                _populate_trench_report_context(
                    context,
                    request,
                    data,
                    property_obj,
                    report_errors=False,
                )
                context['active_calculation_type'] = 'market'
                context['mortgage_form'] = mortgage_form

        elif 'export' in request.POST:
            # РђРЅР°Р»РѕРіРёС‡РЅС‹Рµ РёР·РјРµРЅРµРЅРёСЏ РґР»СЏ Р±Р»РѕРєР° СЌРєСЃРїРѕСЂС‚Р°
            if mortgage_form.is_valid():
                # РџРѕР»СѓС‡Р°РµРј РґР°РЅРЅС‹Рµ РёР· С„РѕСЂРјС‹
                mortgage_data = mortgage_form.cleaned_data
                selected_export_type = request.POST.get('export')
                if selected_export_type not in ('market', 'trench'):
                    selected_export_type = (
                        mortgage_data.get('CALCULATION_TYPE') or 'market'
                    )
                mortgage_data['CALCULATION_TYPE'] = selected_export_type
                context['active_calculation_type'] = selected_export_type

                # РџРѕР»СѓС‡Р°РµРј РІС‹Р±СЂР°РЅРЅС‹Р№ РѕР±СЉРµРєС‚ РЅРµРґРІРёР¶РёРјРѕСЃС‚Рё
                property_obj = mortgage_data['PROPERTY']

                # РџРѕР»СѓС‡Р°РµРј СЃС‚РѕРёРјРѕСЃС‚СЊ РёР· С„РѕСЂРјС‹ Рё РїСЂРµРѕР±СЂР°Р·СѓРµРј РІ float
                property_cost = float(mortgage_data['PROPERTY_COST'])

                (
                    discount_markup_value,
                    discount_markup_rubles,
                    final_property_cost,
                ) = _normalize_discount_markup_values(mortgage_data)
                (
                    initial_payment_percent,
                    initial_payment_rubles,
                ) = _normalize_initial_payment_values(
                    mortgage_data, final_property_cost
                )

                # РЎРѕР·РґР°РµРј СЌРєР·РµРјРїР»СЏСЂ РєР°Р»СЊРєСѓР»СЏС‚РѕСЂР°.
                # Р’СЃРµ Р·РЅР°С‡РµРЅРёСЏ РїСЂРµРѕР±СЂР°Р·СѓРµРј Рє float.
                if mortgage_data.get('CALCULATION_TYPE') == 'trench':
                    trench_mortgage_data, prep_errors = (
                        _prepare_trench_mortgage_data(mortgage_data)
                    )
                    trench_entries, input_rows, trench_errors = (
                        _parse_trench_inputs(
                            post_data=request.POST,
                            trench_count=(
                                trench_mortgage_data['trench_count']
                            ),
                            loan_amount=(
                                trench_mortgage_data['total_loan_amount']
                            ),
                            default_annual_rate=(
                                trench_mortgage_data['annual_rate']
                            ),
                        )
                    )
                    context['trench_input_rows'] = input_rows
                    all_errors = prep_errors + trench_errors
                    if all_errors:
                        context['error_message'] = ' '.join(all_errors)
                        context['active_calculation_type'] = 'trench'
                        return render(
                            request,
                            'mortgage/mortgage_form.html',
                            context,
                        )

                    trench_calculation, calc_errors = (
                        _calculate_trench_mortgage(
                            trench_mortgage_data,
                            trench_entries,
                        )
                    )
                    if calc_errors:
                        context['error_message'] = ' '.join(calc_errors)
                        context['active_calculation_type'] = 'trench'
                        return render(
                            request,
                            'mortgage/mortgage_form.html',
                            context,
                        )
                    trench_calculation['property_obj'] = property_obj
                    return _export_trench_excel(trench_calculation)
                calculator = MortgageCalculator(
                    property_cost=float(final_property_cost),
                    initial_payment_percent=float(initial_payment_percent),
                    initial_payment_date=mortgage_data['INITIAL_PAYMENT_DATE'],
                    mortgage_term=int(mortgage_data['MORTGAGE_TERM']),
                    annual_rate=float(mortgage_data['ANNUAL_RATE']),
                    has_grace_period=mortgage_data['HAS_GRACE_PERIOD']
                    == 'yes',
                    grace_period_term=int(
                        mortgage_data['GRACE_PERIOD_TERM'] or 0
                    ),
                    grace_period_rate=float(
                        mortgage_data['GRACE_PERIOD_RATE'] or 0
                    ),
                )

                # Р’С‹РїРѕР»РЅСЏРµРј СЂР°СЃС‡РµС‚
                result = calculator.calculate()
                payment_schedule = calculator.get_payment_schedule()

                # РЎРѕР·РґР°РµРј Excel-С„Р°Р№Р»
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = (
                    'attachment; filename="mortgage_calculation.xlsx"'
                )

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'РРїРѕС‚РµС‡РЅС‹Р№ СЂР°СЃС‡РµС‚'

                # РЎРѕР·РґР°РµРј СЃС‚РёР»СЊ РґР»СЏ С‡РёСЃРµР» СЃ СЂР°Р·РґРµР»РёС‚РµР»СЏРјРё
                number_style = NamedStyle(name='number_style')
                number_style.number_format = '# ##0.00'
                wb.add_named_style(number_style)

                # РЎРѕР·РґР°РµРј СЃС‚РёР»СЊ РґР»СЏ С†РµР»С‹С… С‡РёСЃРµР»
                integer_style = NamedStyle(name='integer_style')
                integer_style.number_format = '# ##0'
                wb.add_named_style(integer_style)

                # Р—Р°РіРѕР»РѕРІРѕРє
                ws.merge_cells('A1:B1')
                ws['A1'] = 'РРїРѕС‚РµС‡РЅС‹Р№ РєР°Р»СЊРєСѓР»СЏС‚РѕСЂ - СЂРµР·СѓР»СЊС‚Р°С‚С‹ СЂР°СЃС‡РµС‚Р°'
                ws['A1'].font = Font(bold=True, size=14)
                ws['A1'].alignment = Alignment(horizontal='center')

                # Р”Р°РЅРЅС‹Рµ РѕР±СЉРµРєС‚Р°
                ws['A3'] = 'Р”Р°РЅРЅС‹Рµ РѕР±СЉРµРєС‚Р°:'
                ws['A3'].font = Font(bold=True)

                (
                    discount_markup_percent_label,
                    discount_markup_rubles_label,
                ) = _get_discount_markup_labels(
                    mortgage_data['DISCOUNT_MARKUP_TYPE']
                )

                property_data_list = []
                if property_obj is not None:
                    real_estate_complex = (
                        property_obj.building.real_estate_complex
                    )
                    property_data_list.extend(
                        [
                            [
                                'Р“РѕСЂРѕРґ',
                                real_estate_complex.district.city.name,
                            ],
                            ['Р Р°Р№РѕРЅ', real_estate_complex.district.name],
                            [
                                'Р—Р°СЃС‚СЂРѕР№С‰РёРє',
                                real_estate_complex.developer.name,
                            ],
                            ['Р–РёР»РѕР№ РєРѕРјРїР»РµРєСЃ', real_estate_complex.name],
                            ['РљРѕСЂРїСѓСЃ', property_obj.building.number],
                            ['РќРѕРјРµСЂ РєРІР°СЂС‚РёСЂС‹', property_obj.apartment_number],
                            ['РџР»РѕС‰Р°РґСЊ, Рј2', float(property_obj.area)],
                            ['РџР»Р°РЅРёСЂРѕРІРєР°', property_obj.layout.name],
                            ['Р­С‚Р°Р¶', property_obj.floor],
                            ['РћС‚РґРµР»РєР°', property_obj.decoration.name],
                        ]
                    )
                elif mortgage_form.has_manual_property_data():
                    property_data_list.extend(
                        [
                            ['Р“РѕСЂРѕРґ', mortgage_data['OBJECT_CITY'].name],
                            ['Р Р°Р№РѕРЅ', mortgage_data['OBJECT_DISTRICT'].name],
                            [
                                'Р—Р°СЃС‚СЂРѕР№С‰РёРє',
                                mortgage_data['OBJECT_DEVELOPER'].name,
                            ],
                            [
                                'Р–РёР»РѕР№ РєРѕРјРїР»РµРєСЃ',
                                mortgage_data['OBJECT_COMPLEX'].name,
                            ],
                            ['РљРѕСЂРїСѓСЃ', mortgage_data['OBJECT_BUILDING'].number],
                            [
                                'РќРѕРјРµСЂ РєРІР°СЂС‚РёСЂС‹',
                                mortgage_data['OBJECT_APARTMENT_NUMBER'],
                            ],
                            ['РџР»РѕС‰Р°РґСЊ, Рј2', float(mortgage_data['OBJECT_AREA'])],
                            ['РџР»Р°РЅРёСЂРѕРІРєР°', mortgage_data['OBJECT_LAYOUT'].name],
                            ['Р­С‚Р°Р¶', mortgage_data['OBJECT_FLOOR']],
                            ['РћС‚РґРµР»РєР°', mortgage_data['OBJECT_DECORATION'].name],
                        ]
                    )

                property_data_list.extend(
                    [
                        ['Р‘Р°Р·РѕРІР°СЏ СЃС‚РѕРёРјРѕСЃС‚СЊ РѕР±СЉРµРєС‚Р°, СЂСѓР±.', property_cost],
                        [discount_markup_percent_label, discount_markup_value],
                        [discount_markup_rubles_label, discount_markup_rubles],
                        ['РС‚РѕРіРѕРІР°СЏ СЃС‚РѕРёРјРѕСЃС‚СЊ РѕР±СЉРµРєС‚Р°, СЂСѓР±.', final_property_cost],
                    ]
                )

                for i, (param, value) in enumerate(
                    property_data_list, start=4
                ):
                    ws[f'A{i}'] = param
                    cell = ws[f'B{i}']

                    # РџСЂРёРјРµРЅСЏРµРј С„РѕСЂРјР°С‚РёСЂРѕРІР°РЅРёРµ Рє С‡РёСЃР»Р°Рј
                    if isinstance(value, (int, float)):
                        if param == 'Р­С‚Р°Р¶':
                            cell.value = int(value)
                            cell.style = integer_style
                        elif param in [
                            'РџР»РѕС‰Р°РґСЊ, Рј2',
                            discount_markup_percent_label,
                            discount_markup_rubles_label,
                        ]:
                            cell.value = value
                            cell.style = number_style
                        elif param in [
                            'Р‘Р°Р·РѕРІР°СЏ СЃС‚РѕРёРјРѕСЃС‚СЊ РѕР±СЉРµРєС‚Р°, СЂСѓР±.',
                            'РС‚РѕРіРѕРІР°СЏ СЃС‚РѕРёРјРѕСЃС‚СЊ РѕР±СЉРµРєС‚Р°, СЂСѓР±.',
                        ]:
                            cell.value = value
                            cell.style = number_style
                        else:
                            cell.value = value
                    else:
                        cell.value = value

                    # Р’С‹СЂР°РІРЅРёРІР°РЅРёРµ РїРѕ С†РµРЅС‚СЂСѓ РґР»СЏ СЃС‚РѕР»Р±С†Р° B
                    cell.alignment = Alignment(horizontal='center')

                # Р’С…РѕРґРЅС‹Рµ РїР°СЂР°РјРµС‚СЂС‹ РёРїРѕС‚РµРєРё
                start_row = len(property_data_list) + 5
                ws[f'A{start_row}'] = 'РџР°СЂР°РјРµС‚СЂС‹ РёРїРѕС‚РµРєРё:'
                ws[f'A{start_row}'].font = Font(bold=True)

                mortgage_data_list = [
                    ['РџРµСЂРІРѕРЅР°С‡Р°Р»СЊРЅС‹Р№ РІР·РЅРѕСЃ, %', initial_payment_percent],
                    [
                        'РџРµСЂРІРѕРЅР°С‡Р°Р»СЊРЅС‹Р№ РІР·РЅРѕСЃ, СЂСѓР±.',
                        final_property_cost * initial_payment_percent / 100,
                    ],
                    [
                        'Р”Р°С‚Р° РїРµСЂРІРѕРЅР°С‡Р°Р»СЊРЅРѕРіРѕ РІР·РЅРѕСЃР°',
                        mortgage_data['INITIAL_PAYMENT_DATE'].strftime(
                            '%d.%m.%Y'
                        ),
                    ],
                    [
                        'РЎСЂРѕРє РёРїРѕС‚РµРєРё, РіРѕРґС‹',
                        int(mortgage_data['MORTGAGE_TERM_YEARS']),
                    ],
                    [
                        'РЎСЂРѕРє РёРїРѕС‚РµРєРё, РјРµСЃ.',
                        int(mortgage_data['MORTGAGE_TERM']),
                    ],
                    ['Р“РѕРґРѕРІР°СЏ СЃС‚Р°РІРєР°, %', float(mortgage_data['ANNUAL_RATE'])],
                ]

                if mortgage_data['HAS_GRACE_PERIOD'] == 'yes':
                    mortgage_data_list.extend(
                        [
                            [
                                'РЎСЂРѕРє Р»СЊРіРѕС‚РЅРѕРіРѕ РїРµСЂРёРѕРґР°, РіРѕРґС‹',
                                int(
                                    mortgage_data['GRACE_PERIOD_TERM_YEARS']
                                    or 0
                                ),
                            ],
                            [
                                'РЎСЂРѕРє Р»СЊРіРѕС‚РЅРѕРіРѕ РїРµСЂРёРѕРґР°, РјРµСЃ.',
                                int(mortgage_data['GRACE_PERIOD_TERM'] or 0),
                            ],
                            [
                                'Р“РѕРґРѕРІР°СЏ СЃС‚Р°РІРєР° РІ Р»СЊРіРѕС‚РЅС‹Р№ РїРµСЂРёРѕРґ, %',
                                float(mortgage_data['GRACE_PERIOD_RATE']),
                            ],
                        ]
                    )

                for i, (param, value) in enumerate(
                    mortgage_data_list, start=start_row + 1
                ):
                    ws[f'A{i}'] = param
                    cell = ws[f'B{i}']

                    # РџСЂРёРјРµРЅСЏРµРј С„РѕСЂРјР°С‚РёСЂРѕРІР°РЅРёРµ Рє С‡РёСЃР»Р°Рј
                    if isinstance(value, (int, float)):
                        if param in [
                            'РЎСЂРѕРє РёРїРѕС‚РµРєРё, РіРѕРґС‹',
                            'РЎСЂРѕРє РёРїРѕС‚РµРєРё, РјРµСЃ.',
                            'РЎСЂРѕРє Р»СЊРіРѕС‚РЅРѕРіРѕ РїРµСЂРёРѕРґР°, РіРѕРґС‹',
                            'РЎСЂРѕРє Р»СЊРіРѕС‚РЅРѕРіРѕ РїРµСЂРёРѕРґР°, РјРµСЃ.',
                        ]:
                            cell.value = int(value)
                            cell.style = integer_style
                        else:
                            cell.value = value
                            cell.style = number_style
                    else:
                        cell.value = value

                    # Р’С‹СЂР°РІРЅРёРІР°РЅРёРµ РїРѕ С†РµРЅС‚СЂСѓ РґР»СЏ СЃС‚РѕР»Р±С†Р° B
                    cell.alignment = Alignment(horizontal='center')

                # Р РµР·СѓР»СЊС‚Р°С‚С‹ СЂР°СЃС‡РµС‚Р°
                result_start = start_row + len(mortgage_data_list) + 2
                ws[f'A{result_start}'] = 'Р РµР·СѓР»СЊС‚Р°С‚С‹ СЂР°СЃС‡РµС‚Р°:'
                ws[f'A{result_start}'].font = Font(bold=True)

                # Р¤РѕСЂРјР°С‚РёСЂСѓРµРј РґР°С‚Сѓ Р»СЊРіРѕС‚РЅРѕРіРѕ РїРµСЂРёРѕРґР° СЃ РїСЂРѕРІРµСЂРєРѕР№ РЅР° None
                grace_period_end_date_str = ''
                if result['grace_period_end_date']:
                    grace_period_end_date_str = result[
                        'grace_period_end_date'
                    ].strftime('%d.%m.%Y')

                result_data = []

                if mortgage_data['HAS_GRACE_PERIOD'] == 'yes':
                    result_data.extend(
                        [
                            [
                                'Р§РёСЃР»Рѕ РїР»Р°С‚РµР¶РµР№ Р·Р° Р»СЊРіРѕС‚РЅС‹Р№ РїРµСЂРёРѕРґ',
                                result['grace_payments_count'],
                            ],
                            [
                                'Р”Р°С‚Р° РїРѕСЃР»РµРґРЅРµРіРѕ РїР»Р°С‚РµР¶Р° РїРѕ Р»СЊРіРѕС‚РЅРѕРјСѓ РїРµСЂРёРѕРґСѓ',
                                grace_period_end_date_str,
                            ],
                            [
                                (
                                    'РЎСѓРјРјР° РµР¶РµРјРµСЃСЏС‡РЅРѕРіРѕ РїР»Р°С‚РµР¶Р° '
                                    'РІРѕ РІСЂРµРјСЏ Р»СЊРіРѕС‚РЅРѕРіРѕ РїРµСЂРёРѕРґР°, '
                                    'СЂСѓР±.'
                                ),
                                float(result['grace_monthly_payment']),
                            ],
                            [
                                (
                                    'РЎСѓРјРјР° РєСЂРµРґРёС‚Р° РїРѕСЃР»Рµ РѕРєРѕРЅС‡Р°РЅРёСЏ '
                                    'Р»СЊРіРѕС‚РЅРѕРіРѕ РїРµСЂРёРѕРґР°, СЂСѓР±.'
                                ),
                                float(result['loan_after_grace']),
                            ],
                        ]
                    )

                result_data.extend(
                    [
                        [
                            'Р§РёСЃР»Рѕ РїР»Р°С‚РµР¶РµР№ Р·Р° РѕСЃРЅРѕРІРЅРѕР№ РїРµСЂРёРѕРґ',
                            result['main_payments_count'],
                        ],
                        [
                            'Р”Р°С‚Р° РїРѕСЃР»РµРґРЅРµРіРѕ РїР»Р°С‚РµР¶Р° РїРѕ РёРїРѕС‚РµРєРµ',
                            result['mortgage_end_date'].strftime('%d.%m.%Y'),
                        ],
                        [
                            (
                                'РЎСѓРјРјР° РµР¶РµРјРµСЃСЏС‡РЅРѕРіРѕ РїР»Р°С‚РµР¶Р° '
                                'Р·Р° РѕСЃРЅРѕРІРЅРѕР№ РїРµСЂРёРѕРґ, СЂСѓР±.'
                            ),
                            float(result['main_monthly_payment']),
                        ],
                        [
                            'РЎСѓРјРјР° РєСЂРµРґРёС‚Р°, СЂСѓР±.',
                            float(result['total_loan_amount']),
                        ],
                        [
                            'РЎСѓРјРјР° РїРµСЂРµРїР»Р°С‚ РїРѕ РєСЂРµРґРёС‚Сѓ, СЂСѓР±.',
                            float(result['total_overpayment']),
                        ],
                    ]
                )

                for i, (param, value) in enumerate(
                    result_data, start=result_start + 1
                ):
                    ws[f'A{i}'] = param
                    cell = ws[f'B{i}']

                    # РџСЂРёРјРµРЅСЏРµРј С„РѕСЂРјР°С‚РёСЂРѕРІР°РЅРёРµ Рє С‡РёСЃР»Р°Рј
                    if isinstance(value, (int, float)):
                        if param in [
                            'Р§РёСЃР»Рѕ РїР»Р°С‚РµР¶РµР№ Р·Р° Р»СЊРіРѕС‚РЅС‹Р№ РїРµСЂРёРѕРґ',
                            'Р§РёСЃР»Рѕ РїР»Р°С‚РµР¶РµР№ Р·Р° РѕСЃРЅРѕРІРЅРѕР№ РїРµСЂРёРѕРґ',
                        ]:
                            cell.value = int(value)
                            cell.style = integer_style
                        else:
                            cell.value = value
                            cell.style = number_style
                    else:
                        cell.value = value

                    # Р’С‹СЂР°РІРЅРёРІР°РЅРёРµ РїРѕ С†РµРЅС‚СЂСѓ РґР»СЏ СЃС‚РѕР»Р±С†Р° B
                    cell.alignment = Alignment(horizontal='center')

                # Р“СЂР°С„РёРє РїР»Р°С‚РµР¶РµР№
                schedule_start = result_start + len(result_data) + 2
                ws[f'A{schedule_start}'] = 'Р“СЂР°С„РёРє РїР»Р°С‚РµР¶РµР№:'
                ws[f'A{schedule_start}'].font = Font(bold=True)

                headers = [
                    'в„–',
                    'Р”Р°С‚Р° РїР»Р°С‚РµР¶Р°',
                    'РЎСѓРјРјР° РїР»Р°С‚РµР¶Р°, СЂСѓР±.',
                    'Р’ С‚РѕРј С‡РёСЃР»Рµ РїСЂРѕС†РµРЅС‚С‹, СЂСѓР±.',
                    'Р’ С‚РѕРј С‡РёСЃР»Рµ РѕСЃРЅРѕРІРЅРѕР№ РґРѕР»Рі, СЂСѓР±.',
                    'РћСЃС‚Р°С‚РѕРє РґРѕР»РіР°, СЂСѓР±.',
                ]

                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(
                        row=schedule_start + 1, column=col, value=header
                    )
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                for row, payment in enumerate(
                    payment_schedule, start=schedule_start + 2
                ):
                    ws.cell(row=row, column=1, value=payment['payment_number'])
                    ws.cell(
                        row=row,
                        column=2,
                        value=payment['payment_date'].strftime('%d.%m.%Y'),
                    )

                    # РСЃРїРѕР»СЊР·СѓРµРј РёСЃС…РѕРґРЅС‹Рµ С‡РёСЃР»РѕРІС‹Рµ Р·РЅР°С‡РµРЅРёСЏ
                    for col_idx, key in enumerate(
                        [
                            'payment_amount',
                            'interest_amount',
                            'principal_amount',
                            'remaining_debt',
                        ],
                        start=3,
                    ):
                        value = payment[key]
                        # Р•СЃР»Рё Р·РЅР°С‡РµРЅРёРµ - СЃС‚СЂРѕРєР°, РїСЂРµРѕР±СЂР°Р·СѓРµРј РµРіРѕ РІ С‡РёСЃР»Рѕ
                        if isinstance(value, str):
                            numeric_value = float(
                                value.replace(' ', '').replace(',', '.')
                            )
                        else:
                            numeric_value = float(value)
                        cell = ws.cell(
                            row=row, column=col_idx, value=numeric_value
                        )
                        cell.style = number_style
                        cell.alignment = Alignment(horizontal='center')

                # Р¤РѕСЂРјР°С‚РёСЂРѕРІР°РЅРёРµ
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # РЎРѕС…СЂР°РЅСЏРµРј С„Р°Р№Р»
                wb.save(response)
                return response

    return render(request, 'mortgage/mortgage_form.html', context)


def property_cost_api(request, pk):
    """РћРїРёСЃР°РЅРёРµ РјРµС‚РѕРґР° property_cost_api.

    Р’С‹РїРѕР»РЅСЏРµС‚ РїСЂРёРєР»Р°РґРЅСѓСЋ РѕРїРµСЂР°С†РёСЋ С‚РµРєСѓС‰РµРіРѕ РјРѕРґСѓР»СЏ.

    РђСЂРіСѓРјРµРЅС‚С‹:
        request: Р’С…РѕРґРЅРѕР№ РїР°СЂР°РјРµС‚СЂ, РІР»РёСЏСЋС‰РёР№ РЅР° СЂР°Р±РѕС‚Сѓ РјРµС‚РѕРґР°.
        pk: Р’С…РѕРґРЅРѕР№ РїР°СЂР°РјРµС‚СЂ, РІР»РёСЏСЋС‰РёР№ РЅР° СЂР°Р±РѕС‚Сѓ РјРµС‚РѕРґР°.

    Р’РѕР·РІСЂР°С‰Р°РµС‚:
        Any: РўРёРї СЂРµР·СѓР»СЊС‚Р°С‚Р° РѕРїСЂРµРґРµР»СЏРµС‚СЃСЏ РІС‹Р·С‹РІР°СЋС‰РёРј РєРѕРґРѕРј.
    """
    property_obj = get_object_or_404(
        Property.objects.select_related(
            'building__real_estate_complex__developer',
            'building__real_estate_complex__district__city',
            'building',
            'layout',
            'decoration',
        ),
        pk=pk,
    )
    return JsonResponse(_get_property_payload(property_obj))


def calculation_list(request):
    """РЎРїРёСЃРѕРє РІСЃРµС… СЂР°СЃС‡РµС‚РѕРІ"""
    target_customer = _get_target_customer(request)

    if request.method == 'POST' and target_customer is not None:
        selected_ids = request.POST.getlist('calculations')
        calculations = MortgageCalculation.objects.filter(pk__in=selected_ids)

        for calculation in calculations:
            _attach_calculation_to_customer(target_customer, calculation)

        if selected_ids:
            messages.success(
                request,
                'Р’С‹Р±СЂР°РЅРЅС‹Рµ СЂР°СЃС‡РµС‚С‹ РґРѕР±Р°РІР»РµРЅС‹ РІ РєР°СЂС‚РѕС‡РєСѓ РєР»РёРµРЅС‚Р°.',
            )
        else:
            messages.info(request, 'Р Р°СЃС‡РµС‚С‹ РґР»СЏ РґРѕР±Р°РІР»РµРЅРёСЏ РЅРµ РІС‹Р±СЂР°РЅС‹.')
        return redirect('customer:detail', pk=target_customer.pk)

    calculation_filters = get_calculation_filters(request)
    calculation_sort, calculation_order = get_calculation_sort(request)
    calculations = (
        MortgageCalculation.objects.select_related(
            'property',
            'property__layout',
            'property__building',
            'property__building__real_estate_complex',
            'property__building__real_estate_complex__district',
            'property__building__real_estate_complex__district__city',
        )
        .all()
    )
    calculation_cities = get_calculation_city_choices(calculations)
    calculations = apply_calculation_filters(
        annotate_calculation_table_values(calculations), calculation_filters
    )
    calculations = apply_calculation_sort(
        calculations, calculation_sort, calculation_order
    )
    linked_calculation_ids = []
    if target_customer is not None:
        linked_calculation_ids = list(
            target_customer.saved_calculations.values_list('pk', flat=True)
        )
    calculation_filter_reset_url = request.path
    if target_customer is not None:
        calculation_filter_reset_url = (
            f'{request.path}?customer={target_customer.pk}'
        )

    return render(
        request,
        'mortgage/mortgage_list.html',
        {
            'calculations': calculations,
            'target_customer': target_customer,
            'linked_calculation_ids': linked_calculation_ids,
            'calculation_filters': calculation_filters,
            'calculation_cities': calculation_cities,
            'calculation_sort': calculation_sort,
            'calculation_order': calculation_order,
            'calculation_filter_reset_url': calculation_filter_reset_url,
            'calculation_table_headers': build_calculation_table_headers(
                request,
                excluded_fields=('timestamp',),
            ),
        },
    )


@require_POST
def calculation_delete(request, pk):
    """РЈРґР°Р»РµРЅРёРµ СЃРѕС…СЂР°РЅРµРЅРЅРѕРіРѕ РёРїРѕС‚РµС‡РЅРѕРіРѕ СЂР°СЃС‡РµС‚Р°."""
    calculation = get_object_or_404(MortgageCalculation, pk=pk)
    calculation.delete()
    return redirect('mortgage:calculation_list')


def calculation_detail(request, pk):
    """Р”РµС‚Р°Р»СЊРЅР°СЏ РёРЅС„РѕСЂРјР°С†РёСЏ Рѕ СЂР°СЃС‡РµС‚Рµ"""
    calculation = get_object_or_404(
        MortgageCalculation.objects.select_related(
            'property',
            'property__layout',
            'property__building',
            'property__building__real_estate_complex',
            'property__building__real_estate_complex__developer',
            'property__building__real_estate_complex__district',
            'property__building__real_estate_complex__district__city',
            'property__building__real_estate_complex__real_estate_class',
        ),
        pk=pk,
    )
    calculator = MortgageCalculator(
        property_cost=float(calculation.final_property_cost),
        initial_payment_percent=float(calculation.initial_payment_percent),
        initial_payment_date=calculation.initial_payment_date,
        mortgage_term=int(calculation.mortgage_term),
        annual_rate=float(calculation.annual_rate),
        has_grace_period=calculation.has_grace_period,
        grace_period_term=int(calculation.grace_period_term or 0),
        grace_period_rate=float(calculation.grace_period_rate or 0),
    )
    payment_schedule = calculator.get_payment_schedule()

    return render(
        request,
        'mortgage/mortgage_detail.html',
        {
            'calculation': calculation,
            'payment_schedule': payment_schedule,
        },
    )
