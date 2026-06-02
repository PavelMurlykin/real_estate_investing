from dataclasses import dataclass
from decimal import Decimal

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


@dataclass
class MortgageExcelData:
    """Данные для формирования Excel-файла ипотечного расчета."""

    property_obj: object
    mortgage_data: dict
    property_cost: float
    discount_markup_value: float
    discount_markup_rubles: float
    final_property_cost: float
    initial_payment_percent: float
    result: dict
    payment_schedule: list
    has_manual_property_data: bool = False


def get_discount_markup_labels(discount_markup_type):
    """Возвращает подписи для корректировки цены в процентах и рублях."""
    if discount_markup_type == 'discount':
        return 'Скидка, %', 'Скидка, руб.'

    return 'Удорожание, %', 'Удорожание, руб.'


def export_mortgage_excel(excel_data):
    """Формирует HTTP-ответ с Excel-файлом ипотечного расчета."""
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = (
        'attachment; filename="mortgage_calculation.xlsx"'
    )

    workbook = _build_mortgage_workbook(excel_data)
    workbook.save(response)
    return response


def build_saved_mortgage_excel_data(calculation, payment_schedule):
    """Формирует данные экспорта для сохраненного ипотечного расчета."""
    discount_markup_rubles = calculation.base_property_cost
    discount_markup_rubles *= calculation.discount_markup_value / 100
    return MortgageExcelData(
        property_obj=calculation.property,
        mortgage_data={
            'DISCOUNT_MARKUP_TYPE': calculation.discount_markup_type,
            'INITIAL_PAYMENT_DATE': calculation.initial_payment_date,
            'MORTGAGE_TERM_YEARS': calculation.mortgage_term // 12,
            'MORTGAGE_TERM': calculation.mortgage_term,
            'ANNUAL_RATE': calculation.annual_rate,
            'HAS_GRACE_PERIOD': 'yes'
            if calculation.has_grace_period
            else 'no',
            'GRACE_PERIOD_TERM_YEARS': (
                calculation.grace_period_term or 0
            )
            // 12,
            'GRACE_PERIOD_TERM': calculation.grace_period_term or 0,
            'GRACE_PERIOD_RATE': calculation.grace_period_rate or 0,
        },
        property_cost=float(calculation.base_property_cost),
        discount_markup_value=float(calculation.discount_markup_value),
        discount_markup_rubles=float(discount_markup_rubles),
        final_property_cost=float(calculation.final_property_cost),
        initial_payment_percent=float(calculation.initial_payment_percent),
        result={
            'grace_payments_count': calculation.grace_payments_count or 0,
            'grace_period_end_date': calculation.grace_period_end_date,
            'grace_monthly_payment': calculation.grace_monthly_payment or 0,
            'loan_after_grace': calculation.loan_after_grace or 0,
            'main_payments_count': calculation.main_payments_count or 0,
            'mortgage_end_date': calculation.mortgage_end_date,
            'main_monthly_payment': calculation.main_monthly_payment or 0,
            'total_loan_amount': calculation.total_loan_amount or 0,
            'total_overpayment': calculation.total_overpayment or 0,
        },
        payment_schedule=payment_schedule,
    )


def export_saved_mortgage_calculation_excel(calculation, payment_schedule):
    """Формирует Excel-файл для сохраненного ипотечного расчета."""
    excel_data = build_saved_mortgage_excel_data(
        calculation,
        payment_schedule,
    )
    return export_mortgage_excel(excel_data)


def _build_mortgage_workbook(excel_data):
    """Создает книгу Excel с разделами ипотечного расчета."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Ипотечный расчет'

    number_style = NamedStyle(name='number_style')
    number_style.number_format = '# ##0.00'
    workbook.add_named_style(number_style)

    integer_style = NamedStyle(name='integer_style')
    integer_style.number_format = '# ##0'
    workbook.add_named_style(integer_style)

    worksheet.merge_cells('A1:B1')
    worksheet['A1'] = 'Ипотечный калькулятор - результаты расчета'
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A1'].alignment = Alignment(horizontal='center')

    worksheet['A3'] = 'Данные объекта:'
    worksheet['A3'].font = Font(bold=True)

    property_rows = _build_property_rows(excel_data)
    for row_number, (label, value) in enumerate(property_rows, start=4):
        worksheet[f'A{row_number}'] = label
        cell = worksheet[f'B{row_number}']
        _set_value_with_style(
            cell,
            value,
            label,
            number_style,
            integer_style,
            integer_labels={'Этаж'},
            number_labels={
                'Площадь, м2',
                'Скидка, %',
                'Скидка, руб.',
                'Удорожание, %',
                'Удорожание, руб.',
                'Базовая стоимость объекта, руб.',
                'Итоговая стоимость объекта, руб.',
            },
        )

    start_row = len(property_rows) + 5
    worksheet[f'A{start_row}'] = 'Параметры ипотеки:'
    worksheet[f'A{start_row}'].font = Font(bold=True)

    mortgage_rows = _build_mortgage_rows(excel_data)
    for row_number, (label, value) in enumerate(
        mortgage_rows, start=start_row + 1
    ):
        worksheet[f'A{row_number}'] = label
        cell = worksheet[f'B{row_number}']
        _set_value_with_style(
            cell,
            value,
            label,
            number_style,
            integer_style,
            integer_labels={
                'Срок ипотеки, годы',
                'Срок ипотеки, мес.',
                'Срок льготного периода, годы',
                'Срок льготного периода, мес.',
            },
        )

    result_start = start_row + len(mortgage_rows) + 2
    worksheet[f'A{result_start}'] = 'Результаты расчета:'
    worksheet[f'A{result_start}'].font = Font(bold=True)

    result_rows = _build_result_rows(excel_data)
    for row_number, (label, value) in enumerate(
        result_rows, start=result_start + 1
    ):
        worksheet[f'A{row_number}'] = label
        cell = worksheet[f'B{row_number}']
        _set_value_with_style(
            cell,
            value,
            label,
            number_style,
            integer_style,
            integer_labels={
                'Число платежей за льготный период',
                'Число платежей за основной период',
            },
        )

    schedule_start = result_start + len(result_rows) + 2
    worksheet[f'A{schedule_start}'] = 'График платежей:'
    worksheet[f'A{schedule_start}'].font = Font(bold=True)
    _write_payment_schedule(
        worksheet,
        excel_data.payment_schedule,
        schedule_start + 1,
        number_style,
        integer_style,
    )
    _apply_column_widths(worksheet)
    return workbook


def _build_property_rows(excel_data):
    """Возвращает строки раздела с данными объекта."""
    (
        discount_markup_percent_label,
        discount_markup_rubles_label,
    ) = get_discount_markup_labels(
        excel_data.mortgage_data['DISCOUNT_MARKUP_TYPE']
    )

    property_rows = []
    if excel_data.property_obj is not None:
        property_obj = excel_data.property_obj
        real_estate_complex = property_obj.building.real_estate_complex
        property_rows.extend(
            [
                ['Город', real_estate_complex.district.city.name],
                ['Район', real_estate_complex.district.name],
                ['Застройщик', real_estate_complex.developer.name],
                ['Жилой комплекс', real_estate_complex.name],
                ['Корпус', property_obj.building.number],
                ['Номер квартиры', property_obj.apartment_number],
                ['Площадь, м2', float(property_obj.area)],
                ['Планировка', property_obj.layout.name],
                ['Этаж', property_obj.floor],
                ['Отделка', property_obj.decoration.name],
            ]
        )
    elif excel_data.has_manual_property_data:
        mortgage_data = excel_data.mortgage_data
        property_rows.extend(
            [
                ['Город', mortgage_data['OBJECT_CITY'].name],
                ['Район', mortgage_data['OBJECT_DISTRICT'].name],
                ['Застройщик', mortgage_data['OBJECT_DEVELOPER'].name],
                ['Жилой комплекс', mortgage_data['OBJECT_COMPLEX'].name],
                ['Корпус', mortgage_data['OBJECT_BUILDING'].number],
                [
                    'Номер квартиры',
                    mortgage_data['OBJECT_APARTMENT_NUMBER'],
                ],
                ['Площадь, м2', float(mortgage_data['OBJECT_AREA'])],
                ['Планировка', mortgage_data['OBJECT_LAYOUT'].name],
                ['Этаж', mortgage_data['OBJECT_FLOOR']],
                ['Отделка', mortgage_data['OBJECT_DECORATION'].name],
            ]
        )

    property_rows.extend(
        [
            ['Базовая стоимость объекта, руб.', excel_data.property_cost],
            [discount_markup_percent_label, excel_data.discount_markup_value],
            [discount_markup_rubles_label, excel_data.discount_markup_rubles],
            [
                'Итоговая стоимость объекта, руб.',
                excel_data.final_property_cost,
            ],
        ]
    )
    return property_rows


def _build_mortgage_rows(excel_data):
    """Возвращает строки раздела с параметрами ипотеки."""
    mortgage_data = excel_data.mortgage_data
    mortgage_rows = [
        ['Первоначальный взнос, %', excel_data.initial_payment_percent],
        [
            'Первоначальный взнос, руб.',
            excel_data.final_property_cost
            * excel_data.initial_payment_percent
            / 100,
        ],
        [
            'Дата первоначального взноса',
            mortgage_data['INITIAL_PAYMENT_DATE'].strftime('%d.%m.%Y'),
        ],
        [
            'Срок ипотеки, годы',
            int(mortgage_data['MORTGAGE_TERM_YEARS']),
        ],
        ['Срок ипотеки, мес.', int(mortgage_data['MORTGAGE_TERM'])],
        ['Годовая ставка, %', float(mortgage_data['ANNUAL_RATE'])],
    ]

    if _has_grace_period(mortgage_data):
        mortgage_rows.extend(
            [
                [
                    'Срок льготного периода, годы',
                    int(mortgage_data['GRACE_PERIOD_TERM_YEARS'] or 0),
                ],
                [
                    'Срок льготного периода, мес.',
                    int(mortgage_data['GRACE_PERIOD_TERM'] or 0),
                ],
                [
                    'Годовая ставка в льготный период, %',
                    float(mortgage_data['GRACE_PERIOD_RATE']),
                ],
            ]
        )
    return mortgage_rows


def _build_result_rows(excel_data):
    """Возвращает строки раздела с результатами расчета."""
    result = excel_data.result
    mortgage_data = excel_data.mortgage_data
    grace_period_end_date = ''
    if result['grace_period_end_date']:
        grace_period_end_date = result['grace_period_end_date'].strftime(
            '%d.%m.%Y'
        )

    result_rows = []
    if _has_grace_period(mortgage_data):
        result_rows.extend(
            [
                [
                    'Число платежей за льготный период',
                    result['grace_payments_count'],
                ],
                [
                    'Дата последнего платежа по льготному периоду',
                    grace_period_end_date,
                ],
                [
                    (
                        'Сумма ежемесячного платежа '
                        'во время льготного периода, руб.'
                    ),
                    float(result['grace_monthly_payment']),
                ],
                [
                    (
                        'Сумма кредита после окончания '
                        'льготного периода, руб.'
                    ),
                    float(result['loan_after_grace']),
                ],
            ]
        )

    result_rows.extend(
        [
            [
                'Число платежей за основной период',
                result['main_payments_count'],
            ],
            [
                'Дата последнего платежа по ипотеке',
                result['mortgage_end_date'].strftime('%d.%m.%Y'),
            ],
            [
                (
                    'Сумма ежемесячного платежа '
                    'за основной период, руб.'
                ),
                float(result['main_monthly_payment']),
            ],
            ['Сумма кредита, руб.', float(result['total_loan_amount'])],
            [
                'Сумма переплат по кредиту, руб.',
                float(result['total_overpayment']),
            ],
        ]
    )
    return result_rows


def _write_payment_schedule(
    worksheet,
    payment_schedule,
    header_row,
    number_style,
    integer_style,
):
    """Записывает график платежей в лист Excel."""
    headers = [
        '№',
        'Дата платежа',
        'Сумма платежа, руб.',
        'В том числе проценты, руб.',
        'В том числе основной долг, руб.',
        'Остаток долга, руб.',
    ]

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_number, payment in enumerate(
        payment_schedule, start=header_row + 1
    ):
        number_cell = worksheet.cell(
            row=row_number, column=1, value=payment['payment_number']
        )
        number_cell.style = integer_style
        worksheet.cell(
            row=row_number,
            column=2,
            value=payment['payment_date'].strftime('%d.%m.%Y'),
        )

        for column, key in enumerate(
            [
                'payment_amount',
                'interest_amount',
                'principal_amount',
                'remaining_debt',
            ],
            start=3,
        ):
            value = payment[key]
            if isinstance(value, str):
                numeric_value = float(value.replace(' ', '').replace(',', '.'))
            else:
                numeric_value = float(value)
            cell = worksheet.cell(
                row=row_number, column=column, value=numeric_value
            )
            cell.style = number_style
            cell.alignment = Alignment(horizontal='center')


def _set_value_with_style(
    cell,
    value,
    label,
    number_style,
    integer_style,
    integer_labels=None,
    number_labels=None,
):
    """Записывает значение в ячейку и применяет числовой стиль."""
    integer_labels = integer_labels or set()
    number_labels = number_labels or set()
    numeric_types = (int, float, Decimal)

    if isinstance(value, numeric_types):
        if label in integer_labels:
            cell.value = int(value)
            cell.style = integer_style
        elif not number_labels or label in number_labels:
            cell.value = float(value)
            cell.style = number_style
        else:
            cell.value = value
    else:
        cell.value = value
    cell.alignment = Alignment(horizontal='center')


def _apply_column_widths(worksheet):
    """Подбирает ширину столбцов по содержимому."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(
            max_length + 2, 50
        )


def _has_grace_period(mortgage_data):
    """Возвращает признак наличия льготного периода."""
    return mortgage_data.get('HAS_GRACE_PERIOD') in (True, 'yes')
