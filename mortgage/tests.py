from datetime import date
from decimal import Decimal
from io import BytesIO
import re
from zipfile import ZipFile

from openpyxl import load_workbook
from django.test import TestCase
from django.urls import reverse

from bank.models import (
    Bank,
    BankProgram,
    KeyRate,
    MortgageProgram,
    MortgageProgramRegionalCreditLimit,
)
from location.models import City, District, Region
from mortgage.models import MortgageCalculation
from property.models import (
    ApartmentDecoration,
    ApartmentLayout,
    Developer,
    Property,
    RealEstateClass,
    RealEstateComplex,
    RealEstateComplexBuilding,
    RealEstateType,
)
from mortgage.views import _build_saved_trench_calculation_data
from trench_mortgage.models import Trench, TrenchMortgageCalculation


class MortgageCalculatorViewTests(TestCase):
    """Описание класса MortgageCalculatorViewTests.

    Инкапсулирует данные и поведение, необходимые для работы компонента
    в данном модуле.
    """

    def setUp(self):
        """Описание метода setUp.

        Выполняет прикладную операцию текущего модуля.

        Возвращает:
            Any: Тип результата определяется вызывающим кодом.
        """
        self.property = self._create_property(cost=Decimal('5000000.00'))
        self.url = reverse('mortgage:mortgage_calculator')

    def _create_property(
        self,
        cost: Decimal,
        city_name='Город 1',
        complex_name='ЖК Тест',
    ) -> Property:
        """Описание метода _create_property.

        Выполняет прикладную операцию текущего модуля.

        Аргументы:
            cost: Входной параметр, влияющий на работу метода.

        Возвращает:
            Any: Тип результата определяется вызывающим кодом.
        """
        suffix = Region.objects.count() + 1
        region = Region.objects.create(
            name=f'Регион {suffix}', code=f'R{suffix}'
        )
        city = City.objects.create(name=city_name, region=region)
        district = District.objects.create(
            name=f'Район {suffix}', city=city
        )
        developer = Developer.objects.create(name=f'Застройщик {suffix}')
        estate_type = RealEstateType.objects.create(name=f'Квартира {suffix}')
        estate_class = RealEstateClass.objects.create(
            name=f'Комфорт {suffix}', weight=Decimal('1.00')
        )
        complex_obj = RealEstateComplex.objects.create(
            name=complex_name,
            map_link='https://maps.example.com/complex',
            developer=developer,
            district=district,
            real_estate_class=estate_class,
            real_estate_type=estate_type,
        )
        building = RealEstateComplexBuilding.objects.create(
            number='1',
            real_estate_complex=complex_obj,
        )
        layout = ApartmentLayout.objects.create(name=f'1К {suffix}')
        decoration = ApartmentDecoration.objects.create(
            name=f'Без отделки {suffix}'
        )
        return Property.objects.create(
            apartment_number='101',
            building=building,
            decoration=decoration,
            layout=layout,
            area=Decimal('42.00'),
            floor=10,
            property_cost=cost,
        )

    def _create_calculation(
        self,
        property_obj=None,
    ) -> MortgageCalculation:
        if property_obj is None:
            property_obj = self.property

        return MortgageCalculation.objects.create(
            property=property_obj,
            base_property_cost=Decimal('5000000.00'),
            initial_payment_percent=Decimal('20.00'),
            initial_payment_date=date(2026, 1, 1),
            mortgage_term=240,
            annual_rate=Decimal('12.00'),
            has_grace_period=False,
            final_property_cost=Decimal('5000000.00'),
            main_payments_count=240,
            mortgage_end_date=date(2046, 1, 1),
            main_monthly_payment=Decimal('55054.31'),
            total_loan_amount=Decimal('4000000.00'),
            total_overpayment=Decimal('9213034.40'),
        )

    def _create_trench_calculation(
        self,
        property_obj=None,
    ) -> TrenchMortgageCalculation:
        if property_obj is None:
            property_obj = self.property

        calculation = TrenchMortgageCalculation.objects.create(
            property=property_obj,
            base_property_cost=Decimal('5000000.00'),
            discount_markup_type='discount',
            discount_markup_value=Decimal('0.00'),
            final_property_cost=Decimal('5000000.00'),
            initial_payment_percent=Decimal('20.00'),
            initial_payment_date=date(2026, 1, 1),
            mortgage_term=240,
            annual_rate=Decimal('12.00'),
            trench_count=2,
            total_loan_amount=Decimal('4000000.00'),
            total_overpayment=Decimal('8500000.00'),
        )
        Trench.objects.create(
            calculation=calculation,
            trench_number=1,
            trench_date=date(2026, 1, 1),
            trench_percent=Decimal('50.00'),
            trench_amount=Decimal('2000000.00'),
            annual_rate=Decimal('12.00'),
            monthly_payment=Decimal('22021.72'),
            payments_count=12,
            remaining_debt=Decimal('2000000.00'),
        )
        Trench.objects.create(
            calculation=calculation,
            trench_number=2,
            trench_date=date(2027, 1, 1),
            trench_percent=Decimal('50.00'),
            trench_amount=Decimal('2000000.00'),
            annual_rate=Decimal('12.00'),
            monthly_payment=Decimal('44043.44'),
            payments_count=228,
            remaining_debt=Decimal('0.00'),
        )
        return calculation

    def _extract_docx_text(self, content):
        """Возвращает текстовую часть DOCX-файла из HTTP-ответа."""
        document_xml = self._extract_docx_document_xml(content)
        return re.sub(r'<[^>]+>', '', document_xml)

    def _extract_docx_document_xml(self, content):
        """Возвращает XML тела Word-документа."""
        with ZipFile(BytesIO(content)) as archive:
            return archive.read('word/document.xml').decode('utf-8')

    def _extract_docx_table_properties(self, content):
        """Возвращает XML-свойства таблиц Word-документа."""
        document_xml = self._extract_docx_document_xml(content)
        return re.findall(
            r'<w:tblPr>.*?</w:tblPr>',
            document_xml,
            flags=re.DOTALL,
        )

    def _extract_docx_section_properties(self, content):
        """Возвращает XML-свойства секции Word-документа."""
        document_xml = self._extract_docx_document_xml(content)
        return re.findall(
            r'<w:sectPr.*?</w:sectPr>',
            document_xml,
            flags=re.DOTALL,
        )

    def _read_word_template_bytes(self):
        """Возвращает содержимое текущего Word-шаблона отчета."""
        with open(
            'mortgage/word_templates/mortgage_report_template.docx',
            'rb',
        ) as template_file:
            return template_file.read()

    def _extract_docx_relationships(self, content):
        """Возвращает relationships Word-документа."""
        with ZipFile(BytesIO(content)) as archive:
            return archive.read(
                'word/_rels/document.xml.rels'
            ).decode('utf-8')

    def _extract_docx_file_names(self, content):
        """Возвращает имена файлов внутри DOCX."""
        with ZipFile(BytesIO(content)) as archive:
            return archive.namelist()

    def _base_payload(self) -> dict:
        """Описание метода _base_payload.

        Выполняет прикладную операцию текущего модуля.

        Возвращает:
            Any: Тип результата определяется вызывающим кодом.
        """
        return {
            'CALCULATION_TYPE': 'market',
            'PROPERTY': str(self.property.id),
            'DISCOUNT_MARKUP_TYPE': 'discount',
            'DISCOUNT_MARKUP_VALUE': '0',
            'DISCOUNT_MARKUP_RUBLES': '0',
            'DISCOUNT_MARKUP_SOURCE': 'percent',
            'INITIAL_PAYMENT_PERCENT': '20',
            'INITIAL_PAYMENT_RUBLES': '0',
            'INITIAL_PAYMENT_SOURCE': 'percent',
            'INITIAL_PAYMENT_DATE': '01.01.2026',
            'MORTGAGE_TERM_YEARS': '20',
            'MORTGAGE_TERM': '240',
            'ANNUAL_RATE': '12',
            'TRENCH_COUNT': '1',
            'HAS_GRACE_PERIOD': 'no',
            'GRACE_PERIOD_TERM_YEARS': '',
            'GRACE_PERIOD_TERM': '',
            'GRACE_PERIOD_RATE': '',
        }

    def test_get_renders_updated_discount_labels(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Скидка, %')
        self.assertContains(response, 'Скидка, руб.')
        self.assertContains(response, 'initial_payment_source')
        self.assertContains(response, 'discount_markup_percent_lock')
        self.assertContains(response, 'initial_payment_percent_lock')
        self.assertContains(response, 'Срок ипотеки, мес.')
        self.assertContains(response, 'Срок льготного периода, мес.')
        self.assertContains(response, 'id="object-data-collapse"')
        self.assertContains(response, 'calculation-toggle')
        self.assertContains(response, 'aria-label="Показать данные объекта"')
        self.assertContains(response, 'clear-object-data-fields')
        self.assertContains(response, 'Очистить все поля')
        self.assertContains(response, 'property-apartment-menu')
        self.assertContains(
            response,
            'mortgage_form.js?v=20260611-mortgage-programs',
        )
        self.assertContains(response, 'Ипотечная программа')
        self.assertContains(response, 'mortgage-bank-select')
        self.assertContains(response, 'mortgage-program-select')
        self.assertNotContains(
            response,
            'id="mortgage-program-select" name="MORTGAGE_BANK_PROGRAM" '
            'class="form-select" disabled',
        )
        self.assertContains(response, 'mortgage-program-limit-options')
        self.assertContains(response, 'mortgage-program-form-data')
        self.assertContains(response, 'calculation_type')
        self.assertContains(response, 'market-mortgage-pane')
        self.assertContains(response, 'trench-mortgage-pane')
        self.assertContains(response, 'id_TRENCH_COUNT')
        self.assertContains(response, 'trench_amount_source_1')
        self.assertContains(response, 'Стоимость объекта')
        self.assertContains(response, 'Город')
        self.assertContains(response, 'Район')
        self.assertContains(response, 'Жилой комплекс')
        self.assertContains(response, 'Номер квартиры')
        self.assertNotContains(response, 'Объект недвижимости')
        self.assertEqual(
            response.context['property_form_data']['properties'][0]['id'],
            self.property.pk,
        )
        self.assertIn(
            'region_id',
            response.context['property_form_data']['cities'][0],
        )

    def test_calculation_list_is_available_from_calculation_menu(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Сохраненные расчеты ипотеки')
        self.assertContains(response, reverse('mortgage:calculation_list'))
        self.assertNotContains(response, '/trench-mortgage/')

    def test_mortgage_program_block_uses_banks_with_programs(self):
        """Checks mortgage program selector data and credit limits."""
        bank_with_program = Bank.objects.create(
            name='Alpha Bank',
            logo_url='https://img.example/alpha.svg',
        )
        bank_without_program = Bank.objects.create(name='Empty Bank')
        preferential_program = MortgageProgram.objects.create(
            name='Семейная ипотека',
            condition='Льготные условия',
            is_preferential=True,
            credit_limit=Decimal('6000000'),
        )
        MortgageProgramRegionalCreditLimit.objects.create(
            mortgage_program=preferential_program,
            region=self.property.building.real_estate_complex.district.city.region,
            credit_limit=Decimal('12000000'),
        )
        BankProgram.objects.create(
            bank=bank_with_program,
            mortgage_program=preferential_program,
            interest_rate=Decimal('5.90'),
            minimum_initial_payment_percent=Decimal('15.00'),
            maximum_loan_term_years=25,
        )
        KeyRate.objects.create(
            meeting_date=date(2026, 6, 1),
            key_rate=Decimal('16.00'),
        )

        response = self.client.get(self.url)
        program_data = response.context['mortgage_program_form_data']

        self.assertContains(response, 'Alpha Bank')
        self.assertContains(
            response,
            'data-searchable-select-image="https://img.example/alpha.svg"',
        )
        self.assertNotContains(response, 'Empty Bank')
        self.assertEqual(program_data['key_rate'], '16.00')
        self.assertEqual(program_data['banks'][0]['name'], 'Alpha Bank')
        self.assertEqual(
            program_data['banks'][0]['logo_url'],
            'https://img.example/alpha.svg',
        )
        self.assertEqual(
            program_data['programs'][0]['program_name'],
            'Семейная ипотека',
        )
        self.assertEqual(
            program_data['programs'][0]['minimum_initial_payment_percent'],
            '15.00',
        )
        self.assertEqual(program_data['programs'][0]['interest_rate'], '5.90')
        self.assertEqual(
            program_data['programs'][0]['regional_credit_limits'][0][
                'credit_limit'
            ],
            '12000000.00',
        )

    def test_old_trench_mortgage_route_is_removed(self):
        response = self.client.get('/trench-mortgage/')

        self.assertEqual(response.status_code, 404)

    def test_calculation_list_shows_delete_button(self):
        calculation = self._create_calculation()

        response = self.client.get(reverse('mortgage:calculation_list'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Удалить')
        self.assertContains(
            response,
            reverse(
                'mortgage:calculation_delete', kwargs={'pk': calculation.pk}
            ),
        )

    def test_calculation_list_shows_spoiler_actions_and_calculation_details(self):
        calculation = self._create_calculation()
        calculation.base_property_cost = Decimal('5000000.00')
        calculation.final_property_cost = Decimal('4500000.00')
        calculation.discount_markup_type = 'discount'
        calculation.discount_markup_value = Decimal('10.00')
        calculation.initial_payment_percent = Decimal('20.00')
        calculation.mortgage_term = 240
        calculation.annual_rate = Decimal('12.00')
        calculation.has_grace_period = False
        calculation.main_payments_count = 240
        calculation.main_monthly_payment = Decimal('55054.31')
        calculation.save()
        calculation.property.layout_image = 'property/layouts/layout.gif'
        calculation.property.save()

        response = self.client.get(reverse('mortgage:calculation_list'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'css/calculation_table.css')
        self.assertContains(response, 'data-bs-toggle="collapse"')
        self.assertContains(response, 'calculation-toggle')
        self.assertContains(response, 'aria-label="Показать детали расчета"')
        self.assertContains(
            response,
            f'mortgage-calculation-details-{calculation.pk}',
        )
        self.assertContains(
            response,
            reverse(
                'mortgage:calculation_detail', kwargs={'pk': calculation.pk}
            ),
        )
        self.assertContains(
            response,
            reverse(
                'mortgage:calculation_delete', kwargs={'pk': calculation.pk}
            ),
        )
        self.assertContains(response, 'table-bordered')
        self.assertContains(response, 'Стоимость объекта')
        self.assertContains(response, '5 000 000 руб.')
        self.assertContains(
            response,
            '<th scope="row" class="text-nowrap">Скидка</th>',
            html=True,
        )
        self.assertContains(response, 'calculation-detail-table')
        self.assertContains(response, 'style="width: 36ch;"')
        self.assertContains(response, 'style="width: 23ch;"')
        self.assertContains(response, 'calculation-layout-cell')
        self.assertContains(
            response,
            'src="/media/property/layouts/layout.gif"',
        )
        self.assertContains(response, 'data-image-modal="true"')
        self.assertContains(response, 'static/js/image_modal.js')
        self.assertNotContains(
            response,
            'href="/media/property/layouts/layout.gif" target="_blank"',
        )
        self.assertContains(response, 'alt="Планировка"')
        self.assertContains(response, '500 000 руб. (10 %)')
        self.assertContains(response, 'Итоговая стоимость объекта')
        self.assertContains(response, '4 500 000 руб.')
        self.assertContains(response, '900 000 руб. (20 %)')
        self.assertContains(response, '20 лет (240 мес.)')
        self.assertContains(response, 'Сумма ежемесячного платежа')
        self.assertContains(response, '55 054,31 руб.')

    def test_calculation_delete_removes_saved_calculation(self):
        calculation = self._create_calculation()

        response = self.client.post(
            reverse(
                'mortgage:calculation_delete', kwargs={'pk': calculation.pk}
            )
        )

        self.assertRedirects(response, reverse('mortgage:calculation_list'))
        self.assertFalse(
            MortgageCalculation.objects.filter(pk=calculation.pk).exists()
        )

    def test_calculation_list_shows_city_column_and_filter(self):
        calculation = self._create_calculation()

        response = self.client.get(reverse('mortgage:calculation_list'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context['calculation_table_headers'][0]['field'],
            'city',
        )
        self.assertEqual(
            response.context['calculation_table_headers'][1]['field'],
            'object',
        )
        self.assertContains(response, 'Город')
        self.assertNotContains(response, 'Дата расчета')
        self.assertNotContains(
            response,
            calculation.timestamp.strftime('%d.%m.%Y %H:%M'),
        )
        self.assertContains(response, 'name="city"')
        self.assertContains(response, 'Город 1')
        self.assertContains(
            response,
            calculation.property.building.real_estate_complex.name,
        )

    def test_calculation_list_filters_by_city(self):
        moscow_property = self._create_property(
            cost=Decimal('5000000.00'),
            city_name='Москва',
            complex_name='ЖК Москва',
        )
        kazan_property = self._create_property(
            cost=Decimal('5000000.00'),
            city_name='Казань',
            complex_name='ЖК Казань',
        )
        moscow_calculation = self._create_calculation(
            property_obj=moscow_property
        )
        self._create_calculation(property_obj=kazan_property)
        moscow_city = (
            moscow_calculation.property.building.real_estate_complex
            .district.city
        )

        response = self.client.get(
            reverse('mortgage:calculation_list'),
            {'city': moscow_city.pk},
        )
        calculations = list(response.context['calculations'])

        self.assertEqual(response.status_code, 200)
        self.assertEqual(calculations, [moscow_calculation])

    def test_calculation_list_sorts_by_city(self):
        z_property = self._create_property(
            cost=Decimal('5000000.00'),
            city_name='Ярославль',
            complex_name='ЖК Ярославль',
        )
        a_property = self._create_property(
            cost=Decimal('5000000.00'),
            city_name='Астрахань',
            complex_name='ЖК Астрахань',
        )
        self._create_calculation(property_obj=z_property)
        self._create_calculation(property_obj=a_property)

        response = self.client.get(
            reverse('mortgage:calculation_list'),
            {'sort': 'city', 'order': 'asc'},
        )
        cities = [
            calculation.property.building.real_estate_complex.district.city.name
            for calculation in response.context['calculations']
        ]

        self.assertEqual(response.status_code, 200)
        self.assertEqual(cities, ['Астрахань', 'Ярославль'])

    def test_calculation_list_sort_links_use_ajax_without_anchor(self):
        self._create_calculation()

        response = self.client.get(reverse('mortgage:calculation_list'))
        header_urls = [
            header['url']
            for header in response.context['calculation_table_headers']
        ]

        self.assertTrue(header_urls)
        self.assertTrue(
            all('#catalog-results' not in url for url in header_urls)
        )

    def test_calculation_detail_matches_excel_sections(self):
        calculation = self._create_calculation()

        response = self.client.get(
            reverse('mortgage:calculation_detail', kwargs={'pk': calculation.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Данные объекта')
        self.assertContains(response, 'Параметры ипотеки')
        self.assertContains(response, 'Результаты расчета')
        self.assertContains(response, 'График платежей')
        self.assertContains(response, 'table-bordered', count=4)
        self.assertContains(
            response,
            'class="table-responsive mortgage-detail-table-wrapper"',
            count=4,
        )
        self.assertContains(response, 'mortgage-detail-table">', count=4)
        self.assertContains(response, 'width: auto;')
        self.assertContains(response, 'white-space: nowrap;')
        self.assertContains(response, 'Застройщик')
        self.assertContains(response, 'Город')
        self.assertContains(response, 'Название ЖК')
        self.assertContains(response, 'Стоимость объекта, руб.')
        self.assertContains(response, 'Скидка, %')
        self.assertContains(response, 'Итоговая стоимость объекта, руб.')
        self.assertContains(response, 'Первоначальный взнос, %')
        self.assertContains(response, 'Дата первоначального взноса')
        self.assertContains(response, 'Срок ипотеки, годы')
        self.assertContains(response, 'Срок ипотеки, мес.')
        self.assertContains(response, 'Годовая ставка, %')
        self.assertContains(response, 'Число платежей за основной период')
        self.assertContains(response, 'Сумма кредита, руб.')
        self.assertContains(response, 'Сумма переплат по кредиту, руб.')
        self.assertContains(response, 'Дата платежа')
        self.assertContains(response, 'В том числе проценты, руб.')
        self.assertContains(response, 'В том числе основной долг, руб.')
        self.assertContains(response, '01.01.2026')
        self.assertContains(response, '20')
        self.assertContains(response, '240')
        self.assertNotContains(response, 'class="w-50"')
        self.assertNotContains(response, 'Корректировка цены')
        self.assertNotContains(response, 'Наличие льготного периода')
        self.assertNotContains(response, 'Срок льготного периода, годы')

    def test_calculation_detail_has_new_calculation_by_sample_button(self):
        calculation = self._create_calculation()

        response = self.client.get(
            reverse('mortgage:calculation_detail', kwargs={'pk': calculation.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Новый расчет по образцу')
        self.assertContains(
            response,
            f'{reverse("mortgage:mortgage_calculator")}?sample={calculation.pk}',
        )
        self.assertContains(response, 'Сохранить в Excel')
        self.assertContains(response, 'name="export"')
        self.assertContains(response, 'value="market"')
        self.assertContains(response, 'Сохранить в Word')
        self.assertContains(response, 'name="export_word"')
        self.assertContains(response, 'value="market"')

    def test_calculation_detail_export_returns_excel_file(self):
        calculation = self._create_calculation()
        url = reverse(
            'mortgage:calculation_detail', kwargs={'pk': calculation.pk}
        )

        response = self.client.post(url, {'export': 'market'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            (
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )
        self.assertIn(
            'attachment; filename="mortgage_calculation.xlsx"',
            response['Content-Disposition'],
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet.title, 'Ипотечный расчет')
        self.assertEqual(
            worksheet['A1'].value,
            'Ипотечный калькулятор - результаты расчета',
        )
        values = [
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertIn('Данные объекта:', values)
        self.assertIn('Параметры ипотеки:', values)
        self.assertIn('Результаты расчета:', values)
        self.assertIn('График платежей:', values)
        self.assertIn('Сумма платежа, руб.', values)

    def test_calculation_detail_export_returns_word_file(self):
        calculation = self._create_calculation()
        url = reverse(
            'mortgage:calculation_detail', kwargs={'pk': calculation.pk}
        )

        response = self.client.post(url, {'export_word': 'market'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            (
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        self.assertIn(
            'attachment; filename="mortgage_calculation.docx"',
            response['Content-Disposition'],
        )
        document_text = self._extract_docx_text(response.content)
        self.assertIn('Ипотека', document_text)
        self.assertNotIn('Ипотека - результаты расчета', document_text)
        self.assertIn('Локация', document_text)
        self.assertIn('Жилой комплекс', document_text)
        self.assertIn('Параметры объекта недвижимости', document_text)
        self.assertIn('Планировка', document_text)
        self.assertIn('На этаже', document_text)
        self.assertIn('Вид из окна', document_text)
        self.assertIn('ЖК Тест', document_text)
        self.assertIn('Открыть карту ЖК', document_text)
        self.assertIn(
            'Ежемесячный платеж на 240 месяцев (до 01.12.2045)',
            document_text,
        )
        self.assertNotIn('График платежей', document_text)
        self.assertNotIn('Годовая ставка', document_text)
        self.assertNotIn('Скидка', document_text)
        self.assertNotIn('Сумма кредита', document_text)
        self.assertNotIn('Сумма переплаты', document_text)
        self.assertIn(
            'word/media/',
            ' '.join(self._extract_docx_file_names(response.content)),
        )
        relationships = self._extract_docx_relationships(response.content)
        self.assertIn('https://maps.example.com/complex', relationships)
        template_content = self._read_word_template_bytes()
        self.assertEqual(
            self._extract_docx_table_properties(response.content),
            self._extract_docx_table_properties(template_content),
        )
        self.assertEqual(
            self._extract_docx_section_properties(response.content),
            self._extract_docx_section_properties(template_content),
        )

    def test_calculation_detail_word_export_shows_grace_period_payment_row(self):
        calculation = self._create_calculation()
        calculation.has_grace_period = True
        calculation.grace_period_term = 24
        calculation.grace_period_rate = Decimal('6.00')
        calculation.grace_payments_count = 24
        calculation.grace_period_end_date = date(2028, 1, 1)
        calculation.grace_monthly_payment = Decimal('30000.00')
        calculation.loan_after_grace = Decimal('3500000.00')
        calculation.main_payments_count = 216
        calculation.save()
        url = reverse(
            'mortgage:calculation_detail', kwargs={'pk': calculation.pk}
        )

        response = self.client.post(url, {'export_word': 'market'})

        self.assertEqual(response.status_code, 200)
        document_text = self._extract_docx_text(response.content)
        grace_label = (
            'Ежемесячный платеж на 24 месяца (до 01.12.2027)'
        )
        main_label = (
            'Ежемесячный платеж на 216 месяцев (до 01.12.2045)'
        )
        self.assertIn(grace_label, document_text)
        self.assertIn(main_label, document_text)
        self.assertLess(
            document_text.index(grace_label),
            document_text.index(main_label),
        )

    def test_trench_calculation_list_shows_saved_calculations(self):
        calculation = self._create_trench_calculation()

        response = self.client.get(reverse('mortgage:trench_calculation_list'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'История расчетов траншевой ипотеки')
        self.assertContains(response, 'Сохраненные расчеты траншевой ипотеки')
        self.assertContains(response, 'ЖК Тест, кв. 101')
        self.assertContains(response, '5 000 000,00 руб.')
        self.assertContains(response, '8 500 000,00 руб.')
        self.assertContains(
            response,
            reverse(
                'mortgage:trench_calculation_detail',
                kwargs={'pk': calculation.pk},
            ),
        )

    def test_trench_calculation_detail_shows_report_sections(self):
        calculation = self._create_trench_calculation()

        response = self.client.get(
            reverse(
                'mortgage:trench_calculation_detail',
                kwargs={'pk': calculation.pk},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Детали расчета траншевой ипотеки')
        self.assertContains(response, 'Данные объекта')
        self.assertContains(response, 'Параметры траншевой ипотеки')
        self.assertContains(response, 'Транши')
        self.assertContains(response, 'Результаты расчета')
        self.assertContains(response, 'График платежей')
        self.assertContains(response, 'Количество траншей')
        self.assertContains(response, 'Ежемесячный платеж, руб.')
        self.assertContains(response, 'Сохранить в Excel')
        self.assertContains(response, 'name="export"')
        self.assertContains(response, 'value="trench"')
        self.assertContains(response, 'Сохранить в Word')
        self.assertContains(response, 'name="export_word"')
        self.assertContains(response, 'value="trench"')

    def test_saved_trench_calculation_keeps_each_trench_overpayment_positive(
        self,
    ):
        """Calculate saved tranche overpayment for the full repayment term."""
        calculation = self._create_trench_calculation()

        calculation_data = _build_saved_trench_calculation_data(calculation)

        self.assertGreater(calculation_data['trenches'][0]['overpayment'], 0)
        self.assertGreater(calculation_data['trenches'][1]['overpayment'], 0)
        self.assertAlmostEqual(
            calculation_data['trenches'][0]['overpayment'],
            3_285_212.8,
            places=1,
        )

    def test_trench_calculation_detail_export_returns_excel_file(self):
        calculation = self._create_trench_calculation()
        url = reverse(
            'mortgage:trench_calculation_detail',
            kwargs={'pk': calculation.pk},
        )

        response = self.client.post(url, {'export': 'trench'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn(
            'trench_mortgage_calculation.xlsx',
            response['Content-Disposition'],
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet.title, 'Траншевая ипотека')
        values = [
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertIn('Траншевая ипотека - результаты расчета', values)
        self.assertIn('Параметры траншевой ипотеки', values)
        self.assertIn('Транши', values)
        self.assertIn('График платежей', values)

    def test_word_payment_month_label_uses_russian_plural_forms(self):
        payload = self._base_payload()
        payload.update(
            {
                'export_word': 'trench',
                'CALCULATION_TYPE': 'trench',
                'PROPERTY_COST': '5000000',
                'MORTGAGE_TERM': '8',
                'MORTGAGE_TERM_YEARS': '',
                'TRENCH_COUNT': '3',
                'trench_date_1': '2026-01-01',
                'trench_percent_1': '10',
                'trench_amount_1': '',
                'trench_amount_source_1': 'percent',
                'annual_rate_1': '12',
                'trench_date_2': '2026-02-01',
                'trench_percent_2': '10',
                'trench_amount_2': '',
                'trench_amount_source_2': 'percent',
                'annual_rate_2': '12',
                'trench_date_3': '2026-04-01',
                'trench_percent_3': '',
                'trench_amount_3': '',
                'trench_amount_source_3': 'rubles',
                'annual_rate_3': '12',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        document_text = self._extract_docx_text(response.content)
        self.assertIn(
            'Ежемесячный платеж на 1 месяц (до 01.01.2026)',
            document_text,
        )
        self.assertIn(
            'Ежемесячный платеж на 2 месяца (до 01.03.2026)',
            document_text,
        )
        self.assertIn(
            'Ежемесячный платеж на 5 месяцев (до 01.08.2026)',
            document_text,
        )

    def test_trench_calculation_detail_export_returns_word_file(self):
        calculation = self._create_trench_calculation()
        url = reverse(
            'mortgage:trench_calculation_detail',
            kwargs={'pk': calculation.pk},
        )

        response = self.client.post(url, {'export_word': 'trench'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            (
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        self.assertIn(
            'trench_mortgage_calculation.docx',
            response['Content-Disposition'],
        )
        document_text = self._extract_docx_text(response.content)
        self.assertIn('Ипотека траншевая', document_text)
        self.assertNotIn(
            'Ипотека траншевая - результаты расчета',
            document_text,
        )
        self.assertIn('Локация', document_text)
        self.assertIn('Жилой комплекс', document_text)
        self.assertIn('Параметры объекта недвижимости', document_text)
        self.assertIn('Планировка', document_text)
        self.assertIn(
            'Ежемесячный платеж на 12 месяцев (до 01.12.2026)',
            document_text,
        )
        self.assertIn(
            'Ежемесячный платеж на 228 месяцев (до 01.12.2045)',
            document_text,
        )
        self.assertLess(
            document_text.index(
                'Ежемесячный платеж на 12 месяцев (до 01.12.2026)'
            ),
            document_text.index(
                'Ежемесячный платеж на 228 месяцев (до 01.12.2045)'
            ),
        )
        self.assertIn('ЖК Тест', document_text)
        self.assertNotIn('График платежей', document_text)
        self.assertNotIn('Годовая ставка', document_text)
        self.assertNotIn('Количество траншей', document_text)
        self.assertNotIn('Сумма кредита', document_text)
        self.assertNotIn('Сумма переплаты', document_text)

    def test_trench_calculation_delete_removes_saved_calculation(self):
        calculation = self._create_trench_calculation()

        response = self.client.post(
            reverse(
                'mortgage:trench_calculation_delete',
                kwargs={'pk': calculation.pk},
            )
        )

        self.assertRedirects(
            response, reverse('mortgage:trench_calculation_list')
        )
        self.assertFalse(
            TrenchMortgageCalculation.objects.filter(
                pk=calculation.pk
            ).exists()
        )

    def test_mortgage_form_prefills_fields_from_sample_calculation(self):
        calculation = self._create_calculation()
        calculation.base_property_cost = Decimal('6000000.00')
        calculation.discount_markup_type = 'markup'
        calculation.discount_markup_value = Decimal('5.00')
        calculation.final_property_cost = Decimal('6300000.00')
        calculation.initial_payment_percent = Decimal('25.00')
        calculation.initial_payment_date = date(2026, 2, 1)
        calculation.mortgage_term = 180
        calculation.annual_rate = Decimal('11.50')
        calculation.has_grace_period = True
        calculation.grace_period_term = 24
        calculation.grace_period_rate = Decimal('3.50')
        calculation.save()

        response = self.client.get(
            reverse('mortgage:mortgage_calculator'),
            {'sample': calculation.pk},
        )
        form_initial = response.context['mortgage_form'].initial

        self.assertEqual(response.status_code, 200)
        self.assertEqual(form_initial['PROPERTY'], calculation.property_id)
        self.assertEqual(
            form_initial['PROPERTY_COST'],
            Decimal('6000000.00'),
        )
        self.assertEqual(form_initial['DISCOUNT_MARKUP_TYPE'], 'markup')
        self.assertEqual(
            form_initial['DISCOUNT_MARKUP_VALUE'],
            Decimal('5.00'),
        )
        self.assertEqual(
            form_initial['DISCOUNT_MARKUP_RUBLES'],
            Decimal('300000.00'),
        )
        self.assertEqual(form_initial['DISCOUNT_MARKUP_SOURCE'], 'percent')
        self.assertEqual(
            form_initial['INITIAL_PAYMENT_PERCENT'],
            Decimal('25.00'),
        )
        self.assertEqual(
            form_initial['INITIAL_PAYMENT_RUBLES'],
            Decimal('1575000.00'),
        )
        self.assertEqual(form_initial['INITIAL_PAYMENT_SOURCE'], 'percent')
        self.assertEqual(form_initial['INITIAL_PAYMENT_DATE'], date(2026, 2, 1))
        self.assertEqual(form_initial['MORTGAGE_TERM_YEARS'], 15)
        self.assertEqual(form_initial['MORTGAGE_TERM'], 180)
        self.assertEqual(form_initial['ANNUAL_RATE'], Decimal('11.50'))
        self.assertEqual(form_initial['HAS_GRACE_PERIOD'], 'yes')
        self.assertEqual(form_initial['GRACE_PERIOD_TERM_YEARS'], 2)
        self.assertEqual(form_initial['GRACE_PERIOD_TERM'], 24)
        self.assertEqual(form_initial['GRACE_PERIOD_RATE'], Decimal('3.50'))

    def test_calculation_detail_shows_grace_period_rows(self):
        calculation = self._create_calculation()
        calculation.has_grace_period = True
        calculation.grace_period_term = 24
        calculation.grace_period_rate = Decimal('6.00')
        calculation.grace_payments_count = 24
        calculation.grace_period_end_date = date(2028, 1, 1)
        calculation.grace_monthly_payment = Decimal('30000.00')
        calculation.loan_after_grace = Decimal('3500000.00')
        calculation.main_payments_count = 216
        calculation.save()

        response = self.client.get(
            reverse('mortgage:calculation_detail', kwargs={'pk': calculation.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Срок льготного периода, годы')
        self.assertContains(response, 'Срок льготного периода, мес.')
        self.assertContains(response, 'Годовая ставка в льготный период, %')
        self.assertContains(response, 'Число платежей за льготный период')
        self.assertContains(response, 'Дата последнего платежа по льготному периоду')
        self.assertContains(
            response,
            'Сумма ежемесячного платежа во время льготного периода, руб.',
        )

    def test_calculate_uses_property_cost_when_property_cost_is_empty(self):
        """Описание метода
        test_calculate_uses_property_cost_when_property_cost_is_empty.

        Проверяет ожидаемое поведение сценария в рамках автоматического
        теста.

        Возвращает:
            None: Тест завершится ошибкой при нарушении ожидаемого
        поведения.
        """
        payload = self._base_payload()
        payload.update(
            {
                'calculate': '1',
                'PROPERTY_COST': '',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(MortgageCalculation.objects.exists())
        calculation = MortgageCalculation.objects.latest('id')
        self.assertEqual(calculation.base_property_cost, Decimal('5000000.00'))
        self.assertEqual(
            calculation.final_property_cost, Decimal('5000000.00')
        )

    def test_calculate_selects_property_by_building_and_apartment_number(self):
        payload = self._base_payload()
        payload.update(
            {
                'calculate': '1',
                'PROPERTY': '',
                'PROPERTY_COST': '',
                'OBJECT_BUILDING': self.property.building.pk,
                'OBJECT_APARTMENT_NUMBER': self.property.apartment_number,
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        calculation = MortgageCalculation.objects.latest('id')
        self.assertEqual(calculation.property, self.property)
        self.assertEqual(calculation.base_property_cost, Decimal('5000000.00'))

    def test_calculate_without_property_does_not_save_calculation(self):
        payload = self._base_payload()
        payload.update(
            {
                'calculate': '1',
                'PROPERTY': '',
                'PROPERTY_COST': '5000000',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        self.assertIn('result', response.context)
        self.assertFalse(MortgageCalculation.objects.exists())

    def test_trench_calculate_without_property_does_not_save_calculation(self):
        payload = self._base_payload()
        payload.update(
            {
                'calculate': 'trench',
                'CALCULATION_TYPE': 'trench',
                'PROPERTY': '',
                'PROPERTY_COST': '5000000',
                'TRENCH_COUNT': '1',
                'trench_date_1': '2026-01-01',
                'trench_percent_1': '',
                'trench_amount_1': '',
                'trench_amount_source_1': 'rubles',
                'annual_rate_1': '12',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        self.assertIn('trench_result', response.context)
        self.assertEqual(response.context['active_calculation_type'], 'trench')
        self.assertContains(response, 'name="export"')
        self.assertContains(response, 'name="export_word"')
        self.assertFalse(TrenchMortgageCalculation.objects.exists())
        self.assertFalse(MortgageCalculation.objects.exists())

    def test_trench_export_without_property_returns_excel_file(self):
        payload = self._base_payload()
        payload.update(
            {
                'export': 'trench',
                'CALCULATION_TYPE': 'trench',
                'PROPERTY': '',
                'PROPERTY_COST': '5000000',
                'TRENCH_COUNT': '1',
                'trench_date_1': '2026-01-01',
                'trench_percent_1': '',
                'trench_amount_1': '',
                'trench_amount_source_1': 'rubles',
                'annual_rate_1': '12',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn(
            'trench_mortgage_calculation.xlsx',
            response['Content-Disposition'],
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet.title, 'Траншевая ипотека')
        self.assertEqual(
            worksheet['A1'].value,
            'Траншевая ипотека - результаты расчета',
        )
        values = [
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertIn('Данные объекта', values)
        self.assertIn('Параметры траншевой ипотеки', values)
        self.assertIn('Транши', values)
        self.assertIn('График платежей', values)
        self.assertIn('Ежемесячный платеж, руб.', values)

    def test_trench_calculate_saves_trench_calculation_from_mortgage_form(self):
        payload = self._base_payload()
        payload.update(
            {
                'calculate': 'trench',
                'CALCULATION_TYPE': 'trench',
                'TRENCH_COUNT': '2',
                'trench_date_1': '2026-01-01',
                'trench_percent_1': '2.50',
                'trench_amount_1': '100000',
                'trench_amount_source_1': 'rubles',
                'annual_rate_1': '12',
                'trench_date_2': '2026-07-01',
                'trench_percent_2': '',
                'trench_amount_2': '',
                'trench_amount_source_2': 'rubles',
                'annual_rate_2': '12',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        self.assertIn('trench_result', response.context)
        self.assertIn('result', response.context)
        self.assertFalse(MortgageCalculation.objects.exists())
        calculation = TrenchMortgageCalculation.objects.get()
        trenches = list(Trench.objects.order_by('trench_number'))
        self.assertEqual(calculation.property, self.property)
        self.assertEqual(calculation.trench_count, 2)
        self.assertEqual(
            calculation.total_loan_amount,
            Decimal('4000000.00'),
        )
        self.assertEqual(len(trenches), 2)
        self.assertEqual(trenches[0].trench_amount, Decimal('100000.00'))
        self.assertEqual(trenches[1].trench_amount, Decimal('3900000.00'))

    def test_market_calculate_keeps_trench_report_when_trench_data_is_present(self):
        payload = self._base_payload()
        payload.update(
            {
                'calculate': 'market',
                'TRENCH_COUNT': '2',
                'trench_date_1': '2026-01-01',
                'trench_percent_1': '2.50',
                'trench_amount_1': '100000',
                'trench_amount_source_1': 'rubles',
                'annual_rate_1': '12',
                'trench_date_2': '2026-07-01',
                'trench_percent_2': '',
                'trench_amount_2': '',
                'trench_amount_source_2': 'rubles',
                'annual_rate_2': '12',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['active_calculation_type'], 'market')
        self.assertIn('result', response.context)
        self.assertIn('trench_result', response.context)
        self.assertTrue(MortgageCalculation.objects.exists())
        self.assertFalse(TrenchMortgageCalculation.objects.exists())

    def test_calculate_with_manual_object_creates_property_and_calculation(self):
        payload = self._base_payload()
        payload.update(
            {
                'calculate': '1',
                'PROPERTY': '',
                'PROPERTY_COST': '7000000',
                'OBJECT_CITY': (
                    self.property.building.real_estate_complex.district.city.pk
                ),
                'OBJECT_DISTRICT': (
                    self.property.building.real_estate_complex.district.pk
                ),
                'OBJECT_DEVELOPER': (
                    self.property.building.real_estate_complex.developer.pk
                ),
                'OBJECT_COMPLEX': (
                    self.property.building.real_estate_complex.pk
                ),
                'OBJECT_BUILDING': self.property.building.pk,
                'OBJECT_APARTMENT_NUMBER': '202',
                'OBJECT_AREA': '55.50',
                'OBJECT_LAYOUT': self.property.layout.pk,
                'OBJECT_FLOOR': '12',
                'OBJECT_DECORATION': self.property.decoration.pk,
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        created_property = Property.objects.get(apartment_number='202')
        self.assertEqual(created_property.property_cost, Decimal('7000000'))
        calculation = MortgageCalculation.objects.latest('id')
        self.assertEqual(calculation.property, created_property)
        self.assertEqual(calculation.base_property_cost, Decimal('7000000'))

    def test_calculate_keeps_overridden_cost_only_in_calculation(self):
        """Описание метода
        test_calculate_keeps_overridden_cost_only_in_calculation.

        Проверяет ожидаемое поведение сценария в рамках автоматического
        теста.

        Возвращает:
            None: Тест завершится ошибкой при нарушении ожидаемого
        поведения.
        """
        payload = self._base_payload()
        payload.update(
            {
                'calculate': '1',
                'PROPERTY_COST': '6000000',
                'DISCOUNT_MARKUP_TYPE': 'discount',
                'DISCOUNT_MARKUP_VALUE': '10',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        calculation = MortgageCalculation.objects.latest('id')
        self.assertEqual(calculation.base_property_cost, Decimal('6000000'))
        self.assertEqual(calculation.final_property_cost, Decimal('5400000'))

        self.property.refresh_from_db()
        self.assertEqual(self.property.property_cost, Decimal('5000000.00'))

    def test_calculate_supports_discount_markup_in_rubles(self):
        payload = self._base_payload()
        payload.update(
            {
                'calculate': '1',
                'PROPERTY_COST': '5000000',
                'DISCOUNT_MARKUP_VALUE': '0',
                'DISCOUNT_MARKUP_RUBLES': '500000',
                'DISCOUNT_MARKUP_SOURCE': 'rubles',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        calculation = MortgageCalculation.objects.latest('id')
        self.assertEqual(calculation.discount_markup_value, Decimal('10'))
        self.assertEqual(calculation.final_property_cost, Decimal('4500000'))
        self.assertContains(response, 'Скидка, руб.')

    def test_calculate_supports_initial_payment_locked_in_rubles(self):
        payload = self._base_payload()
        payload.update(
            {
                'calculate': '1',
                'PROPERTY_COST': '5000000',
                'INITIAL_PAYMENT_PERCENT': '20',
                'INITIAL_PAYMENT_RUBLES': '1500000',
                'INITIAL_PAYMENT_SOURCE': 'rubles',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        calculation = MortgageCalculation.objects.latest('id')
        self.assertEqual(calculation.initial_payment_percent, Decimal('30'))

    def test_calculate_saves_mortgage_and_grace_terms_in_months(self):
        payload = self._base_payload()
        payload.update(
            {
                'calculate': '1',
                'PROPERTY_COST': '5000000',
                'MORTGAGE_TERM_YEARS': '2',
                'MORTGAGE_TERM': '30',
                'HAS_GRACE_PERIOD': 'yes',
                'GRACE_PERIOD_TERM_YEARS': '1',
                'GRACE_PERIOD_TERM': '14',
                'GRACE_PERIOD_RATE': '6',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        calculation = MortgageCalculation.objects.latest('id')
        self.assertEqual(calculation.mortgage_term, 30)
        self.assertEqual(calculation.grace_period_term, 14)
        self.assertEqual(calculation.grace_payments_count, 14)
        self.assertEqual(calculation.main_payments_count, 16)

    def test_export_returns_excel_file(self):
        """Описание метода test_export_returns_excel_file.

        Проверяет ожидаемое поведение сценария в рамках автоматического
        теста.

        Возвращает:
            None: Тест завершится ошибкой при нарушении ожидаемого
        поведения.
        """
        payload = self._base_payload()
        payload.update(
            {
                'export': '1',
                'PROPERTY_COST': '5100000',
                'DISCOUNT_MARKUP_TYPE': 'markup',
                'DISCOUNT_MARKUP_VALUE': '5',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn(
            'attachment; filename="mortgage_calculation.xlsx"',
            response['Content-Disposition'],
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet.title, 'Ипотечный расчет')
        self.assertEqual(
            worksheet['A1'].value,
            'Ипотечный калькулятор - результаты расчета',
        )
        self.assertEqual(worksheet['A3'].value, 'Данные объекта:')
        values = [
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertIn('Параметры ипотеки:', values)
        self.assertIn('Результаты расчета:', values)
        self.assertIn('График платежей:', values)
        self.assertIn('Сумма платежа, руб.', values)
        self.assertNotIn('Корректировка цены', values)
        self.assertNotIn('Наличие льготного периода', values)

    def test_export_returns_word_file(self):
        payload = self._base_payload()
        payload.update(
            {
                'export_word': 'market',
                'PROPERTY_COST': '5100000',
                'DISCOUNT_MARKUP_TYPE': 'markup',
                'DISCOUNT_MARKUP_VALUE': '5',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            (
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        self.assertIn(
            'mortgage_calculation.docx',
            response['Content-Disposition'],
        )
        document_text = self._extract_docx_text(response.content)
        self.assertIn('Локация', document_text)
        self.assertIn('Жилой комплекс', document_text)
        self.assertIn('Параметры объекта недвижимости', document_text)
        self.assertIn('Ипотека', document_text)
        self.assertNotIn('Ипотека - результаты расчета', document_text)
        self.assertIn('ЖК Тест', document_text)
        self.assertIn(
            'Ежемесячный платеж на 240 месяцев (до 01.12.2045)',
            document_text,
        )
        self.assertNotIn('График платежей', document_text)
        self.assertNotIn('Годовая ставка', document_text)
        self.assertNotIn('Удорожание, %', document_text)
        self.assertNotIn('Сумма кредита', document_text)
        self.assertNotIn('Сумма переплаты', document_text)

    def test_trench_export_returns_word_file(self):
        payload = self._base_payload()
        payload.update(
            {
                'export_word': 'trench',
                'CALCULATION_TYPE': 'trench',
                'PROPERTY_COST': '5000000',
                'TRENCH_COUNT': '1',
                'trench_date_1': '2026-01-01',
                'trench_percent_1': '',
                'trench_amount_1': '',
                'trench_amount_source_1': 'rubles',
                'annual_rate_1': '12',
            }
        )

        response = self.client.post(self.url, payload)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            (
                'application/vnd.openxmlformats-officedocument.'
                'wordprocessingml.document'
            ),
        )
        self.assertIn(
            'trench_mortgage_calculation.docx',
            response['Content-Disposition'],
        )
        document_text = self._extract_docx_text(response.content)
        self.assertIn('Ипотека траншевая', document_text)
        self.assertNotIn(
            'Ипотека траншевая - результаты расчета',
            document_text,
        )
        self.assertIn('Локация', document_text)
        self.assertIn('Жилой комплекс', document_text)
        self.assertIn('Параметры объекта недвижимости', document_text)
        self.assertIn(
            'Ежемесячный платеж на 240 месяцев (до 01.12.2045)',
            document_text,
        )
        self.assertNotIn('График платежей', document_text)
        self.assertNotIn('Годовая ставка', document_text)
        self.assertNotIn('Количество траншей', document_text)
        self.assertNotIn('Сумма кредита', document_text)
        self.assertNotIn('Сумма переплаты', document_text)

    def test_property_cost_api_returns_cost_from_database(self):
        """Описание метода test_property_cost_api_returns_cost_from_database.

        Проверяет ожидаемое поведение сценария в рамках автоматического
        теста.

        Возвращает:
            None: Тест завершится ошибкой при нарушении ожидаемого
        поведения.
        """
        url = reverse(
            'mortgage:property_cost_api', kwargs={'pk': self.property.pk}
        )

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['property_cost'], '5000000.00')
        self.assertEqual(
            response.json()['city_id'],
            self.property.building.real_estate_complex.district.city.pk,
        )
        self.assertEqual(
            response.json()['building_id'],
            self.property.building.pk,
        )
        self.assertEqual(response.json()['apartment_number'], '101')
