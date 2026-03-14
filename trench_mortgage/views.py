from datetime import datetime
from decimal import Decimal, InvalidOperation
from math import pow

import openpyxl
from dateutil.relativedelta import relativedelta
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from mortgage.utils import format_currency
from property.models import Property

from .forms import TrenchMortgageForm
from .models import Trench, TrenchMortgageCalculation

MAX_TRENCH_COUNT = 5


def trench_mortgage_calculator(request):
    """Описание метода trench_mortgage_calculator.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        request: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    initial = _build_initial_data(request.GET.get('property_id'))

    if request.method == 'POST':
        form_data = request.POST.copy()
        _hydrate_property_cost_fields(form_data)
        form = TrenchMortgageForm(form_data)
    else:
        form = TrenchMortgageForm(initial=initial)

    trench_count = _resolve_trench_count(form)
    default_rate = _resolve_default_rate(form)
    context = {
        'form': form,
        'trench_count': trench_count,
        'trench_input_rows': _build_trench_input_rows(
            trench_count=trench_count,
            post_data=form.data if form.is_bound else None,
            default_annual_rate=default_rate,
        ),
        'final_property_cost': _calculate_display_final_cost(
            form.data if form.is_bound else form.initial
        ),
    }

    if request.method != 'POST' or not form.is_valid():
        return render(request, 'trench_mortgage/calculator.html', context)

    mortgage_data, prep_errors = _prepare_mortgage_data(form.cleaned_data)
    trench_entries, input_rows, trench_errors = _parse_trench_inputs(
        post_data=request.POST,
        trench_count=mortgage_data['trench_count'],
        loan_amount=mortgage_data['total_loan_amount'],
        default_annual_rate=mortgage_data['annual_rate'],
    )
    context['trench_input_rows'] = input_rows
    context['final_property_cost'] = format_currency(
        mortgage_data['final_property_cost']
    )

    all_errors = prep_errors + trench_errors
    if all_errors:
        context['error_message'] = ' '.join(all_errors)
        return render(request, 'trench_mortgage/calculator.html', context)

    calculation, calc_errors = _calculate_trench_mortgage(
        mortgage_data, trench_entries
    )
    if calc_errors:
        context['error_message'] = ' '.join(calc_errors)
        return render(request, 'trench_mortgage/calculator.html', context)

    if 'export' in request.POST:
        return _export_trench_excel(calculation)

    _save_trench_calculation(calculation)
    context['result'] = _format_result(calculation)
    context['payment_schedule'] = _format_payment_schedule(
        calculation['payment_schedule']
    )
    return render(request, 'trench_mortgage/calculator.html', context)


def _build_initial_data(raw_property_id):
    """Описание метода _build_initial_data.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        raw_property_id: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    initial = {}
    if not (raw_property_id and str(raw_property_id).isdigit()):
        return initial

    property_obj = Property.objects.filter(pk=int(raw_property_id)).first()
    if not property_obj:
        return initial

    initial['PROPERTY'] = property_obj.id
    initial['PROPERTY_COST'] = property_obj.property_cost
    return initial


def _hydrate_property_cost_fields(form_data):
    """Описание метода _hydrate_property_cost_fields.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        form_data: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    selected_id = form_data.get('PROPERTY')
    if not selected_id:
        return

    selected_property = Property.objects.filter(id=selected_id).first()
    if not selected_property:
        return

    if not form_data.get('PROPERTY_COST'):
        form_data['PROPERTY_COST'] = str(selected_property.property_cost)


def _resolve_trench_count(form):
    """Описание метода _resolve_trench_count.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        form: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    raw_count = None
    if form.is_bound:
        raw_count = form.data.get('TRENCH_COUNT')
    elif 'TRENCH_COUNT' in form.initial:
        raw_count = form.initial.get('TRENCH_COUNT')
    else:
        raw_count = form.fields['TRENCH_COUNT'].initial

    try:
        count = int(raw_count or 1)
    except (TypeError, ValueError):
        count = 1

    return max(1, min(MAX_TRENCH_COUNT, count))


def _resolve_default_rate(form):
    """Описание метода _resolve_default_rate.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        form: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    if form.is_bound:
        return str(form.data.get('ANNUAL_RATE', '')).strip()
    return str(form.initial.get('ANNUAL_RATE', '')).strip()


def _build_trench_input_rows(
    trench_count, post_data=None, default_annual_rate=''
):
    """Описание метода _build_trench_input_rows.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        trench_count: Входной параметр, влияющий на работу метода.
        post_data: Входной параметр, влияющий на работу метода.
        default_annual_rate: Входной параметр, влияющий на работу
    метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    rows = []
    default_rate_value = str(default_annual_rate or '').strip()

    for idx in range(1, MAX_TRENCH_COUNT + 1):
        if post_data:
            trench_date = str(post_data.get(f'trench_date_{idx}', '')).strip()
            trench_percent = str(
                post_data.get(f'trench_percent_{idx}', '')
            ).strip()
            trench_amount = str(
                post_data.get(f'trench_amount_{idx}', '')
            ).strip()
            annual_rate = str(post_data.get(f'annual_rate_{idx}', '')).strip()
        else:
            trench_date = ''
            trench_percent = ''
            trench_amount = ''
            annual_rate = ''

        if not annual_rate:
            annual_rate = default_rate_value

        rows.append(
            {
                'number': idx,
                'trench_date': trench_date,
                'trench_percent': trench_percent,
                'trench_amount': trench_amount,
                'annual_rate': annual_rate,
                'is_active': idx <= trench_count,
                'is_last': idx == trench_count,
            }
        )

    return rows


def _calculate_display_final_cost(data):
    """Описание метода _calculate_display_final_cost.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        data: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    if not data:
        return ''
    if not data.get('PROPERTY_COST'):
        return ''

    try:
        property_cost = float(data.get('PROPERTY_COST') or 0)
        discount_value = float(data.get('DISCOUNT_MARKUP_VALUE') or 0)
    except (TypeError, ValueError):
        return ''

    discount_type = data.get('DISCOUNT_MARKUP_TYPE') or 'discount'
    if discount_type == 'discount':
        final_cost = property_cost * (1 - discount_value / 100)
    else:
        final_cost = property_cost * (1 + discount_value / 100)

    return format_currency(final_cost)


def _prepare_mortgage_data(cleaned_data):
    """Описание метода _prepare_mortgage_data.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        cleaned_data: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    errors = []
    property_obj = cleaned_data['PROPERTY']

    property_cost = float(cleaned_data['PROPERTY_COST'])
    base_property_cost = float(cleaned_data['PROPERTY_COST'])
    discount_markup_value = float(
        cleaned_data.get('DISCOUNT_MARKUP_VALUE', 0) or 0
    )
    discount_markup_type = cleaned_data['DISCOUNT_MARKUP_TYPE']

    if discount_markup_type == 'discount':
        final_property_cost = property_cost * (1 - discount_markup_value / 100)
    else:
        final_property_cost = property_cost * (1 + discount_markup_value / 100)

    initial_payment_percent = float(
        cleaned_data.get('INITIAL_PAYMENT_PERCENT', 0) or 0
    )
    initial_payment_rubles = float(
        cleaned_data.get('INITIAL_PAYMENT_RUBLES', 0) or 0
    )

    if (
        initial_payment_rubles
        and not initial_payment_percent
        and final_property_cost > 0
    ):
        initial_payment_percent = (
            initial_payment_rubles / final_property_cost
        ) * 100

    if (
        initial_payment_percent
        and not initial_payment_rubles
        and final_property_cost > 0
    ):
        initial_payment_rubles = (
            final_property_cost * initial_payment_percent / 100
        )

    if initial_payment_percent and initial_payment_rubles:
        initial_payment_rubles = (
            final_property_cost * initial_payment_percent / 100
        )

    total_loan_amount = final_property_cost - initial_payment_rubles
    if total_loan_amount <= 0:
        errors.append(
            (
                'Сумма кредита должна быть больше 0 после вычета '
                'первоначального взноса.'
            )
        )

    return {
        'property_obj': property_obj,
        'property_cost': property_cost,
        'base_property_cost': base_property_cost,
        'discount_markup_type': discount_markup_type,
        'discount_markup_value': discount_markup_value,
        'final_property_cost': final_property_cost,
        'initial_payment_percent': initial_payment_percent,
        'initial_payment_rubles': initial_payment_rubles,
        'initial_payment_date': cleaned_data['INITIAL_PAYMENT_DATE'],
        'mortgage_term': int(cleaned_data['MORTGAGE_TERM']),
        'annual_rate': float(cleaned_data['ANNUAL_RATE']),
        'trench_count': int(cleaned_data['TRENCH_COUNT']),
        'total_loan_amount': total_loan_amount,
    }, errors


def _parse_decimal_value(raw_value):
    """Описание метода _parse_decimal_value.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        raw_value: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    return Decimal(str(raw_value).replace(',', '.'))


def _parse_trench_inputs(
    post_data, trench_count, loan_amount, default_annual_rate
):
    """Описание метода _parse_trench_inputs.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        post_data: Входной параметр, влияющий на работу метода.
        trench_count: Входной параметр, влияющий на работу метода.
        loan_amount: Входной параметр, влияющий на работу метода.
        default_annual_rate: Входной параметр, влияющий на работу
    метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    errors = []
    input_rows = _build_trench_input_rows(
        trench_count, post_data, default_annual_rate
    )
    entries = []
    percent_sum = Decimal('0')
    loan_amount_decimal = Decimal(str(loan_amount))
    default_rate_decimal = Decimal(str(default_annual_rate))

    for idx in range(1, trench_count + 1):
        trench_date_raw = str(post_data.get(f'trench_date_{idx}', '')).strip()
        trench_percent_raw = str(
            post_data.get(f'trench_percent_{idx}', '')
        ).strip()
        trench_amount_raw = str(
            post_data.get(f'trench_amount_{idx}', '')
        ).strip()
        annual_rate_raw = str(post_data.get(f'annual_rate_{idx}', '')).strip()
        row_data = input_rows[idx - 1]
        row_data['trench_date'] = trench_date_raw
        row_data['trench_percent'] = trench_percent_raw
        row_data['trench_amount'] = trench_amount_raw
        row_data['annual_rate'] = (
            annual_rate_raw or f'{default_rate_decimal:.2f}'
        )

        if not trench_date_raw:
            errors.append(f'Не заполнена дата транша №{idx}.')
            continue

        try:
            trench_date = datetime.strptime(trench_date_raw, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f'Неверный формат даты транша №{idx}.')
            continue

        if annual_rate_raw:
            try:
                annual_rate = _parse_decimal_value(annual_rate_raw)
            except (InvalidOperation, ValueError):
                errors.append(f'Неверная ставка транша №{idx}.')
                continue
        else:
            annual_rate = default_rate_decimal

        if annual_rate < 0:
            errors.append(f'Ставка транша №{idx} не может быть отрицательной.')
            continue

        trench_percent = None
        trench_amount = None
        if idx < trench_count:
            if not trench_percent_raw and not trench_amount_raw:
                errors.append(
                    f'Не заполнена сумма транша №{idx} (в % или руб.).'
                )
                continue

            if trench_percent_raw:
                try:
                    trench_percent = _parse_decimal_value(trench_percent_raw)
                except (InvalidOperation, ValueError):
                    errors.append(f'Неверный процент транша №{idx}.')
                    continue

            if trench_amount_raw:
                try:
                    trench_amount = _parse_decimal_value(trench_amount_raw)
                except (InvalidOperation, ValueError):
                    errors.append(f'Неверная сумма транша №{idx} в рублях.')
                    continue

            if trench_percent is not None and trench_percent <= 0:
                errors.append(f'Процент транша №{idx} должен быть больше 0.')
                continue
            if trench_amount is not None and trench_amount <= 0:
                errors.append(
                    f'Сумма транша №{idx} в рублях должна быть больше 0.'
                )
                continue

            if trench_percent is not None:
                computed_percent = trench_percent
                computed_amount = (
                    loan_amount_decimal * computed_percent / Decimal('100')
                )
            else:
                computed_amount = trench_amount
                if loan_amount_decimal <= 0:
                    errors.append('Сумма кредита должна быть больше 0.')
                    continue
                computed_percent = (
                    computed_amount / loan_amount_decimal
                ) * Decimal('100')

            percent_sum += computed_percent
            trench_percent = computed_percent.quantize(Decimal('0.01'))
            trench_amount = computed_amount.quantize(Decimal('0.01'))
        entries.append(
            {
                'number': idx,
                'trench_date': trench_date,
                'trench_percent': trench_percent,
                'trench_amount': trench_amount,
                'annual_rate': annual_rate.quantize(Decimal('0.01')),
            }
        )

    if errors:
        return [], input_rows, errors

    if len(entries) != trench_count:
        return [], input_rows, ['Некорректные данные траншей.']

    if trench_count == 1:
        last_percent = Decimal('100')
    else:
        last_percent = Decimal('100') - percent_sum

    if last_percent <= 0:
        return (
            [],
            input_rows,
            ['Сумма процентов траншей превышает или равна 100%.'],
        )

    last_amount = (
        loan_amount_decimal * last_percent / Decimal('100')
    ).quantize(Decimal('0.01'))
    entries[-1]['trench_percent'] = last_percent.quantize(Decimal('0.01'))
    entries[-1]['trench_amount'] = last_amount

    for idx, entry in enumerate(entries):
        input_rows[idx]['trench_percent'] = '{:.2f}'.format(
            entry['trench_percent']
        )
        input_rows[idx]['trench_amount'] = '{:.2f}'.format(
            entry['trench_amount']
        )
        input_rows[idx]['annual_rate'] = '{:.2f}'.format(entry['annual_rate'])

    for prev, current in zip(entries, entries[1:]):
        if current['trench_date'] < prev['trench_date']:
            return [], input_rows, ['Даты траншей должны идти по возрастанию.']

    return entries, input_rows, []


def _calculate_trench_mortgage(mortgage_data, trench_entries):
    """Описание метода _calculate_trench_mortgage.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        mortgage_data: Входной параметр, влияющий на работу метода.
        trench_entries: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    errors = []

    loan_amount = float(mortgage_data['total_loan_amount'])
    initial_payment_date = mortgage_data['initial_payment_date']
    mortgage_end_date = initial_payment_date + relativedelta(
        years=mortgage_data['mortgage_term']
    )

    trenches_result = []
    total_overpayment = 0.0
    cumulative_amount = 0.0

    for idx, trench in enumerate(trench_entries):
        trench_date = trench['trench_date']
        trench_number = trench['number']
        if trench_date < initial_payment_date:
            errors.append(
                (
                    f'Дата транша №{trench_number} не может быть '
                    'раньше даты первоначального взноса.'
                )
            )
            continue

        months_remaining = _calculate_months_remaining(
            trench_date, mortgage_end_date
        )
        if months_remaining <= 0:
            errors.append(
                (
                    f'Для транша №{trench_number} нет месяцев '
                    'до конца срока ипотеки.'
                )
            )
            continue

        trench_percent = float(trench['trench_percent'])
        trench_amount = float(trench['trench_amount'])
        annual_rate = float(trench['annual_rate'])
        monthly_rate = annual_rate / 100 / 12

        if monthly_rate > 0:
            factor = pow(1 + monthly_rate, months_remaining)
            monthly_payment = (
                trench_amount * monthly_rate * factor / (factor - 1)
            )
        else:
            monthly_payment = trench_amount / months_remaining

        overpayment = (monthly_payment * months_remaining) - trench_amount
        if idx + 1 < len(trench_entries):
            period_end_date = trench_entries[idx + 1]['trench_date']
        else:
            period_end_date = mortgage_end_date
        payments_count = _calculate_months_remaining(
            trench_date, period_end_date
        )

        cumulative_amount += trench_amount
        remaining_debt = max(loan_amount - cumulative_amount, 0)
        total_overpayment += overpayment

        trenches_result.append(
            {
                'number': trench['number'],
                'date': trench_date,
                'percent': trench_percent,
                'amount': trench_amount,
                'annual_rate': annual_rate,
                'monthly_payment': monthly_payment,
                'payments_count': payments_count,
                'remaining_debt': remaining_debt,
                'overpayment': overpayment,
            }
        )

    if errors:
        return None, errors

    calculation = {
        'property_obj': mortgage_data['property_obj'],
        'property_cost': mortgage_data['property_cost'],
        'base_property_cost': mortgage_data['base_property_cost'],
        'discount_markup_type': mortgage_data['discount_markup_type'],
        'discount_markup_value': mortgage_data['discount_markup_value'],
        'final_property_cost': mortgage_data['final_property_cost'],
        'initial_payment_percent': mortgage_data['initial_payment_percent'],
        'initial_payment': mortgage_data['initial_payment_rubles'],
        'initial_payment_date': mortgage_data['initial_payment_date'],
        'mortgage_term': mortgage_data['mortgage_term'],
        'mortgage_end_date': mortgage_end_date,
        'trench_count': mortgage_data['trench_count'],
        'total_loan_amount': mortgage_data['total_loan_amount'],
        'total_overpayment': total_overpayment,
        'trenches': trenches_result,
        'payment_schedule': _build_trench_payment_schedule(
            trenches_result, mortgage_end_date
        ),
    }

    return calculation, []


def _calculate_months_remaining(start_date, end_date):
    """Описание метода _calculate_months_remaining.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        start_date: Входной параметр, влияющий на работу метода.
        end_date: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    months = (end_date.year - start_date.year) * 12 + (
        end_date.month - start_date.month
    )
    if end_date.day < start_date.day:
        months -= 1
    return months


def _build_trench_payment_schedule(trenches, mortgage_end_date):
    """Описание метода _build_trench_payment_schedule.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        trenches: Входной параметр, влияющий на работу метода.
        mortgage_end_date: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    if not trenches:
        return []

    trench_states = []
    for trench in trenches:
        months_to_end = _calculate_months_remaining(
            trench['date'], mortgage_end_date
        )
        if months_to_end <= 0:
            continue

        trench_states.append(
            {
                'start_date': trench['date'],
                'months_left': months_to_end,
                'monthly_rate': trench['annual_rate'] / 100 / 12,
                'monthly_payment': trench['monthly_payment'],
                'balance': trench['amount'],
            }
        )

    if not trench_states:
        return []

    first_payment_date = min(state['start_date'] for state in trench_states)
    total_schedule_months = _calculate_months_remaining(
        first_payment_date, mortgage_end_date
    )
    payment_schedule = []

    for month_index in range(total_schedule_months):
        payment_date = first_payment_date + relativedelta(months=month_index)
        payment_amount = 0.0
        interest_amount = 0.0
        principal_amount = 0.0

        for state in trench_states:
            if payment_date < state['start_date'] or state['months_left'] <= 0:
                continue

            trench_interest = (
                state['balance'] * state['monthly_rate']
                if state['monthly_rate'] > 0
                else 0.0
            )
            trench_principal = max(
                state['monthly_payment'] - trench_interest, 0.0
            )

            if (
                state['months_left'] == 1
                or trench_principal >= state['balance']
            ):
                trench_principal = state['balance']

            trench_payment = trench_interest + trench_principal
            state['balance'] = max(state['balance'] - trench_principal, 0.0)
            state['months_left'] -= 1

            payment_amount += trench_payment
            interest_amount += trench_interest
            principal_amount += trench_principal

        remaining_debt = sum(
            max(state['balance'], 0.0)
            for state in trench_states
            if state['start_date'] <= payment_date
        )
        payment_schedule.append(
            {
                'payment_number': month_index + 1,
                'payment_date': payment_date,
                'payment_amount': round(payment_amount, 2),
                'interest_amount': round(interest_amount, 2),
                'principal_amount': round(principal_amount, 2),
                'remaining_debt': round(remaining_debt, 2),
            }
        )

    return payment_schedule


def _format_result(calculation):
    """Описание метода _format_result.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        calculation: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    formatted = {
        'final_property_cost': format_currency(
            calculation['final_property_cost']
        ),
        'initial_payment': format_currency(calculation['initial_payment']),
        'initial_payment_date': calculation['initial_payment_date'].strftime(
            '%d.%m.%Y'
        ),
        'total_loan_amount': format_currency(calculation['total_loan_amount']),
        'total_overpayment': format_currency(calculation['total_overpayment']),
        'trenches': [],
    }

    for trench in calculation['trenches']:
        formatted['trenches'].append(
            {
                'number': trench['number'],
                'date': trench['date'].strftime('%d.%m.%Y'),
                'percent': '{:.2f}'.format(trench['percent']),
                'amount': format_currency(trench['amount']),
                'annual_rate': '{:.2f}'.format(trench['annual_rate']),
                'monthly_payment': format_currency(trench['monthly_payment']),
                'payments_count': trench['payments_count'],
                'remaining_debt': format_currency(trench['remaining_debt']),
                'overpayment': format_currency(trench['overpayment']),
            }
        )

    return formatted


def _format_payment_schedule(payment_schedule):
    """Описание метода _format_payment_schedule.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        payment_schedule: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    formatted_schedule = []
    for payment in payment_schedule:
        formatted_schedule.append(
            {
                'payment_number': payment['payment_number'],
                'payment_date': payment['payment_date'].strftime('%d.%m.%Y'),
                'payment_amount': format_currency(payment['payment_amount']),
                'interest_amount': format_currency(payment['interest_amount']),
                'principal_amount': format_currency(
                    payment['principal_amount']
                ),
                'remaining_debt': format_currency(payment['remaining_debt']),
            }
        )

    return formatted_schedule


def _save_trench_calculation(calculation):
    """Описание метода _save_trench_calculation.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        calculation: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    calc_obj = TrenchMortgageCalculation.objects.create(
        property=calculation['property_obj'],
        final_property_cost=Decimal(str(calculation['final_property_cost'])),
        initial_payment_percent=Decimal(
            str(calculation['initial_payment_percent'])
        ),
        initial_payment_date=calculation['initial_payment_date'],
        mortgage_term=calculation['mortgage_term'],
        trench_count=calculation['trench_count'],
        total_loan_amount=Decimal(str(calculation['total_loan_amount'])),
        total_overpayment=Decimal(str(calculation['total_overpayment'])),
    )

    trench_objects = []
    for trench in calculation['trenches']:
        trench_objects.append(
            Trench(
                calculation=calc_obj,
                trench_number=trench['number'],
                trench_date=trench['date'],
                trench_percent=Decimal(str(trench['percent'])),
                trench_amount=Decimal(str(trench['amount'])),
                annual_rate=Decimal(str(trench['annual_rate'])),
                monthly_payment=Decimal(str(trench['monthly_payment'])),
                payments_count=trench['payments_count'],
                remaining_debt=Decimal(str(trench['remaining_debt'])),
            )
        )

    if trench_objects:
        Trench.objects.bulk_create(trench_objects)


def _export_trench_excel(calculation):
    """Описание метода _export_trench_excel.

    Выполняет прикладную операцию текущего модуля.

    Аргументы:
        calculation: Входной параметр, влияющий на работу метода.

    Возвращает:
        Any: Тип результата определяется вызывающим кодом.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="trench_mortgage_calculation.xlsx"'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Траншевая ипотека'

    number_style = NamedStyle(name='number_style')
    number_style.number_format = '# ##0.00'
    wb.add_named_style(number_style)

    integer_style = NamedStyle(name='integer_style')
    integer_style.number_format = '# ##0'
    wb.add_named_style(integer_style)

    ws.merge_cells('A1:B1')
    ws['A1'] = 'Траншевая ипотека - результаты расчета'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    property_obj = calculation['property_obj']
    property_rows = [
        [
            'Застройщик',
            property_obj.building.real_estate_complex.developer.name,
        ],
        [
            'Город',
            property_obj.building.real_estate_complex.district.city.name,
        ],
        ['Название ЖК', property_obj.building.real_estate_complex.name],
        ['Корпус', property_obj.building.number],
        ['№ квартиры', property_obj.apartment_number],
        ['Площадь', float(property_obj.area)],
        ['Этаж', property_obj.floor],
        ['Стоимость объекта, руб.', calculation['property_cost']],
        [
            'Тип изменения цены',
            'Скидка'
            if calculation['discount_markup_type'] == 'discount'
            else 'Удорожание',
        ],
        ['Значение, %', calculation['discount_markup_value']],
        [
            'Итоговая стоимость объекта, руб.',
            calculation['final_property_cost'],
        ],
    ]

    row = 3
    ws[f'A{row}'] = 'Данные объекта'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    for label, value in property_rows:
        ws[f'A{row}'] = label
        cell = ws[f'B{row}']
        cell.value = value
        if isinstance(value, int):
            cell.style = integer_style
        elif isinstance(value, float):
            cell.style = number_style
        cell.alignment = Alignment(horizontal='center')
        row += 1

    row += 1
    ws[f'A{row}'] = 'Параметры траншевой ипотеки'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    mortgage_rows = [
        ['Первоначальный взнос, %', calculation['initial_payment_percent']],
        ['Первоначальный взнос, руб.', calculation['initial_payment']],
        [
            'Дата первоначального взноса',
            calculation['initial_payment_date'].strftime('%d.%m.%Y'),
        ],
        ['Срок кредита, лет', calculation['mortgage_term']],
        ['Количество траншей', calculation['trench_count']],
    ]

    for label, value in mortgage_rows:
        ws[f'A{row}'] = label
        cell = ws[f'B{row}']
        cell.value = value
        if isinstance(value, int):
            cell.style = integer_style
        elif isinstance(value, float):
            cell.style = number_style
        cell.alignment = Alignment(horizontal='center')
        row += 1

    row += 1
    ws[f'A{row}'] = 'Транши'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    headers = [
        '№',
        'Дата',
        'Сумма, %',
        'Сумма, руб.',
        'Ставка, %',
        'Ежемесячный платеж, руб.',
        'Число платежей',
        'Остаток долга, руб.',
        'Переплата, руб.',
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row += 1
    for trench in calculation['trenches']:
        ws.cell(
            row=row, column=1, value=trench['number']
        ).style = integer_style
        ws.cell(row=row, column=2, value=trench['date'].strftime('%d.%m.%Y'))

        values = [
            trench['percent'],
            trench['amount'],
            trench['annual_rate'],
            trench['monthly_payment'],
            trench['payments_count'],
            trench['remaining_debt'],
            trench['overpayment'],
        ]

        for col_offset, value in enumerate(values, start=3):
            cell = ws.cell(row=row, column=col_offset, value=float(value))
            cell.style = integer_style if col_offset == 7 else number_style
            cell.alignment = Alignment(horizontal='center')

        row += 1

    row += 1
    ws[f'A{row}'] = 'Итоги'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    totals = [
        ['Сумма кредита, руб.', calculation['total_loan_amount']],
        ['Сумма переплаты, руб.', calculation['total_overpayment']],
    ]

    for label, value in totals:
        ws[f'A{row}'] = label
        cell = ws[f'B{row}']
        cell.value = float(value)
        cell.style = number_style
        cell.alignment = Alignment(horizontal='center')
        row += 1

    if calculation['payment_schedule']:
        row += 1
        ws[f'A{row}'] = 'График платежей'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        schedule_headers = [
            '№',
            'Дата платежа',
            'Сумма платежа, руб.',
            'Проценты, руб.',
            'Погашение основного долга, руб.',
            'Остаток долга, руб.',
        ]
        for col, header in enumerate(schedule_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        row += 1
        for payment in calculation['payment_schedule']:
            ws.cell(
                row=row, column=1, value=payment['payment_number']
            ).style = integer_style
            ws.cell(
                row=row,
                column=2,
                value=payment['payment_date'].strftime('%d.%m.%Y'),
            )

            schedule_values = [
                payment['payment_amount'],
                payment['interest_amount'],
                payment['principal_amount'],
                payment['remaining_debt'],
            ]
            for col_offset, value in enumerate(schedule_values, start=3):
                cell = ws.cell(row=row, column=col_offset, value=float(value))
                cell.style = number_style
                cell.alignment = Alignment(horizontal='center')

            row += 1

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(response)
    return response
