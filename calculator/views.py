from django.shortcuts import render
from django.http import HttpResponse
from .forms import MortgageForm
from .models import MortgageCalculation
from .mortgage_calculator import MortgageCalculator
from .utils import format_currency
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
import decimal


def mortgage_calculator(request):
    context = {'form': MortgageForm()}

    if request.method == 'POST' and 'calculate' in request.POST:
        form = MortgageForm(request.POST)
        if form.is_valid():
            # Получаем данные из формы
            data = form.cleaned_data

            # Создаем экземпляр калькулятора
            calculator = MortgageCalculator(
                property_cost=float(data['PROPERTY_COST']),
                initial_payment_percent=float(data['INITIAL_PAYMENT_PERCENT']),
                initial_payment_date=data['INITIAL_PAYMENT_DATE'],
                mortgage_term=data['MORTGAGE_TERM'],
                annual_rate=float(data['ANNUAL_RATE']),
                has_grace_period=data['HAS_GRACE_PERIOD'] == 'да',
                grace_period_term=data['GRACE_PERIOD_TERM'] or 0,
                grace_period_rate=float(data['GRACE_PERIOD_RATE'] or 0)
            )

            # Выполняем расчет
            result = calculator.calculate()

            # Форматируем числовые значения для отображения
            formatted_result = {}
            for key, value in result.items():
                if isinstance(value, (int, float, decimal.Decimal)):
                    formatted_result[key] = format_currency(value)
                else:
                    formatted_result[key] = value

            # Получаем график платежей
            payment_schedule = calculator.get_payment_schedule()

            # Форматируем числовые значения в графике платежей
            for payment in payment_schedule:
                for key in ['payment_amount', 'interest_amount', 'principal_amount', 'remaining_debt']:
                    if key in payment:
                        payment[key] = format_currency(payment[key])

            # Сохраняем расчет в контекст
            context['result'] = formatted_result
            context['payment_schedule'] = payment_schedule
            context['form'] = form

            # Сохраняем в базу данных
            calculation = MortgageCalculation(
                property_cost=data['PROPERTY_COST'],
                initial_payment_percent=data['INITIAL_PAYMENT_PERCENT'],
                initial_payment_date=data['INITIAL_PAYMENT_DATE'],
                mortgage_term=data['MORTGAGE_TERM'],
                annual_rate=data['ANNUAL_RATE'],
                has_grace_period=data['HAS_GRACE_PERIOD'] == 'да',
                grace_period_term=data['GRACE_PERIOD_TERM'],
                grace_period_rate=data['GRACE_PERIOD_RATE'],
                # Результаты
                grace_payments_count=result['grace_payments_count'],
                grace_period_end_date=result['grace_period_end_date'],
                grace_monthly_payment=result['grace_monthly_payment'],
                loan_after_grace=result['loan_after_grace'],
                main_payments_count=result['main_payments_count'],
                mortgage_end_date=result['mortgage_end_date'],
                main_monthly_payment=result['main_monthly_payment'],
                total_loan_amount=result['total_loan_amount'],
                total_overpayment=result['total_overpayment']
            )
            calculation.save()

    elif request.method == 'POST' and 'export' in request.POST:
        form = MortgageForm(request.POST)
        if form.is_valid():
            # Получаем данные из формы
            data = form.cleaned_data

            # Создаем экземпляр калькулятора
            calculator = MortgageCalculator(
                property_cost=float(data['PROPERTY_COST']),
                initial_payment_percent=float(data['INITIAL_PAYMENT_PERCENT']),
                initial_payment_date=data['INITIAL_PAYMENT_DATE'],
                mortgage_term=data['MORTGAGE_TERM'],
                annual_rate=float(data['ANNUAL_RATE']),
                has_grace_period=data['HAS_GRACE_PERIOD'] == 'да',
                grace_period_term=data['GRACE_PERIOD_TERM'] or 0,
                grace_period_rate=float(data['GRACE_PERIOD_RATE'] or 0)
            )

            # Выполняем расчет
            result = calculator.calculate()
            payment_schedule = calculator.get_payment_schedule()

            # Создаем Excel-файл
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="mortgage_calculation.xlsx"'

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ипотечный расчет"

            # Создаем стиль для чисел с разделителями
            number_style = NamedStyle(name="number_style")
            number_style.number_format = '# ##0.00'
            wb.add_named_style(number_style)

            # Заголовок
            ws.merge_cells('A1:B1')
            ws['A1'] = 'Ипотечный калькулятор - результаты расчета'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Входные параметры
            ws['A3'] = 'Входные параметры:'
            ws['A3'].font = Font(bold=True)

            input_data = [
                ['Стоимость объекта, руб.', data['PROPERTY_COST']],
                ['Первоначальный взнос, %', data['INITIAL_PAYMENT_PERCENT']],
                ['Дата первоначального взноса', data['INITIAL_PAYMENT_DATE'].strftime('%d.%m.%Y')],
                ['Срок ипотеки, годы', data['MORTGAGE_TERM']],
                ['Годовая ставка, %', data['ANNUAL_RATE']],
                ['Наличие льготного периода', 'Да' if data['HAS_GRACE_PERIOD'] == 'да' else 'Нет'],
            ]

            if data['HAS_GRACE_PERIOD'] == 'да':
                input_data.extend([
                    ['Срок льготного периода, годы', data['GRACE_PERIOD_TERM']],
                    ['Годовая ставка в льготный период, %', data['GRACE_PERIOD_RATE']]
                ])

            for i, (param, value) in enumerate(input_data, start=4):
                ws[f'A{i}'] = param
                cell = ws[f'B{i}']

                # Применяем форматирование к числам
                if isinstance(value, (int, float, decimal.Decimal)):
                    cell.value = value
                    cell.style = number_style
                else:
                    cell.value = value

            # Результаты расчета
            start_row = len(input_data) + 5
            ws[f'A{start_row}'] = 'Результаты расчета:'
            ws[f'A{start_row}'].font = Font(bold=True)

            result_data = [
                ['Число платежей за льготный период', result['grace_payments_count']],
                ['Дата последнего платежа по льготному периоду', result['grace_period_end_date'].strftime('%d.%m.%Y')],
                ['Сумма ежемесячного платежа во время льготного периода, руб.', result['grace_monthly_payment']],
                ['Сумма кредита после окончания льготного периода, руб.', result['loan_after_grace']],
                ['Число платежей за основной период', result['main_payments_count']],
                ['Дата последнего платежа по ипотеке', result['mortgage_end_date'].strftime('%d.%m.%Y')],
                ['Сумма ежемесячного платежа за основной период, руб.', result['main_monthly_payment']],
                ['Сумма кредита, руб.', result['total_loan_amount']],
                ['Сумма переплат по кредиту, руб.', result['total_overpayment']],
            ]

            for i, (param, value) in enumerate(result_data, start=start_row + 1):
                ws[f'A{i}'] = param
                cell = ws[f'B{i}']

                # Применяем форматирование к числам
                if isinstance(value, (int, float, decimal.Decimal)):
                    cell.value = value
                    cell.style = number_style
                else:
                    cell.value = value

            # График платежей
            schedule_start = start_row + len(result_data) + 2
            ws[f'A{schedule_start}'] = 'График платешений:'
            ws[f'A{schedule_start}'].font = Font(bold=True)

            headers = ['№', 'Дата платежа', 'Сумма платежа, руб.', 'В том числе проценты, руб.',
                       'В том числе основной долг, руб.', 'Остаток долга, руб.']

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=schedule_start + 1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for row, payment in enumerate(payment_schedule, start=schedule_start + 2):
                ws.cell(row=row, column=1, value=payment['payment_number'])
                ws.cell(row=row, column=2, value=payment['payment_date'].strftime('%d.%m.%Y'))

                # Форматируем числовые значения
                for col_idx, key in enumerate(
                        ['payment_amount', 'interest_amount', 'principal_amount', 'remaining_debt'], start=3):
                    cell = ws.cell(row=row, column=col_idx, value=payment[key])
                    cell.style = number_style

            # Форматирование
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем файл
            wb.save(response)
            return response

    return render(request, 'calculator/index.html', context)